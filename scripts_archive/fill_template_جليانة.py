import pandas as pd
import sys
import os
from openpyxl import load_workbook

def normalize_name(name):
    if not isinstance(name, str): return ""
    name = str(name)
    if name.lower() in ['الاسم', 'اسم الموظف', 'الرقم الوظيفي', 'nan', 'none']: 
        return ""
    return " ".join(name.strip().split())

def map_relationship(rel):
    rel = str(rel).strip()
    if rel in ['موظف', 'nan', 'Principal', 'نفسه'] or not rel or rel == 'None' or rel == 'الصلة':
        return ""
    
    mapping = {
        'ابن': 'SON',
        'ابنة': 'DAUGHTER',
        'بنت': 'DAUGHTER',
        'زوج': 'HUSBAND',
        'زوجة': 'WIFE',
        'زوجه': 'WIFE',
        'أب': 'FATHER',
        'اب': 'FATHER',
        'أم': 'MOTHER',
        'ام': 'MOTHER',
        'أخ': 'BROTHER',
        'اخ': 'BROTHER',
        'أخت': 'SISTER',
        'اخت': 'SISTER'
    }
    return mapping.get(rel, rel)

try:
    # 1. Process source data
    source_file = r'd:\tba_waad_system-main\tba_waad_system-main\القائمة_الموحدة_جليانة_جاهزة_للرفع_دقيقة_100.xlsx'
    df = pd.read_excel(source_file)
    
    df = df.dropna(how='all')
    df['الاسم'] = df['الاسم'].apply(normalize_name)
    df = df[df['الاسم'] != '']
    df['اسم_الموظف'] = df['اسم_الموظف'].apply(normalize_name)
    df['الرقم_الوطني'] = df['الرقم_الوطني'].fillna('').astype(str).str.replace('.0', '', regex=False).str.strip()
    
    dependent_rels = ['ابن', 'ابنة', 'بنت', 'زوج', 'زوجة', 'زوجه', 'أب', 'اب', 'أم', 'ام', 'أخ', 'اخ', 'أخت', 'اخت']
    df['is_principal'] = df.apply(lambda row: str(row['الصلة']).strip() not in dependent_rels, axis=1)
    
    df['has_national_id'] = (df['الرقم_الوطني'] != '') & (df['الرقم_الوطني'] != 'nan') & (df['الرقم_الوطني'] != '0')
    
    with_id = df[df['has_national_id']]
    without_id = df[~df['has_national_id']]
    
    with_id = with_id.drop_duplicates(subset=['الرقم_الوطني'])
    without_id = without_id.drop_duplicates(subset=['الاسم', 'اسم_الموظف', 'الصلة'])
    
    df_clean = pd.concat([with_id, without_id])
    df_clean = df_clean.drop_duplicates(subset=['الاسم', 'اسم_الموظف'])
    
    # 2. Prepare data for template
    # Template columns: full_name, employer, principal_card_number, relationship
    final_data = []
    for _, row in df_clean.iterrows():
        final_data.append({
            'full_name': row['الاسم'],
            'employer': 'المنطقة الحرة جليانة' if row['is_principal'] else '', # Matched from lookup
            'principal_card_number': '',
            'relationship': map_relationship(row['الصلة']),
            'SORT_EMP': row['اسم_الموظف'],
            'SORT_TYPE': 1 if row['is_principal'] else 0
        })
    
    final_df = pd.DataFrame(final_data)
    final_df = final_df.sort_values(by=['SORT_EMP', 'SORT_TYPE'], ascending=[True, False])
    
    # Remove sort columns
    export_df = final_df[['full_name', 'employer', 'principal_card_number', 'relationship']]
    
    # 3. Load Template and Fill
    template_path = r'd:\tba_waad_system-main\tba_waad_system-main\members_template (3).xlsx'
    output_path = r'd:\tba_waad_system-main\tba_waad_system-main\جليانة.xlsx'
    
    wb = load_workbook(template_path)
    if 'Data' not in wb.sheetnames:
        print("Error: 'Data' sheet not found in template")
        sys.exit(1)
        
    ws = wb['Data']
    
    # Clear existing data except header
    # Row 1 is header, Row 2 is example.
    # Openpyxl is 1-indexed. Row 1 = row 0 in pandas.
    # ws.delete_rows(2, ws.max_row) # This might mess up formatting if any.
    
    # Better to just write starting from row 2
    row_idx = 2
    for _, series in export_df.iterrows():
        ws.cell(row=row_idx, column=1, value=series['full_name'])
        ws.cell(row=row_idx, column=2, value=series['employer'])
        ws.cell(row=row_idx, column=3, value=series['principal_card_number'])
        ws.cell(row=row_idx, column=4, value=series['relationship'])
        row_idx += 1
        
    wb.save(output_path)
    print(f"Data filled successfully. Final rows: {len(export_df)}")
    print(f"Saved to: {output_path}")

except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)
