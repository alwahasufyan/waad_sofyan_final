
import pandas as pd
import sys
from openpyxl import load_workbook

def normalize_name(name):
    if not isinstance(name, str): return ""
    name = str(name).strip()
    return " ".join(name.split())

def map_relationship(rel):
    rel = str(rel).strip()
    if not rel or rel == 'nan' or rel in ['موظف', 'Principal', 'نفسه']: return ""
    mapping = {
        'ابن': 'SON', 'ابنة': 'DAUGHTER', 'بنت': 'DAUGHTER',
        'زوج': 'HUSBAND', 'زوجة': 'WIFE', 'زوجه': 'WIFE',
        'أب': 'FATHER', 'اب': 'FATHER', 'أم': 'MOTHER', 'ام': 'MOTHER',
        'أخ': 'BROTHER', 'اخ': 'BROTHER', 'أخت': 'SISTER', 'اخت': 'SISTER'
    }
    return mapping.get(rel, rel)

try:
    source_file = r'd:\tba_waad_system-main\tba_waad_system-main\القائمة_الموحدة_جليانة_جاهزة_للرفع_دقيقة_100.xlsx'
    df = pd.read_excel(source_file)
    
    df['n1'] = df['الاسم'].apply(normalize_name)
    df['n2'] = df['اسم_الموظف'].apply(normalize_name)
    df['card'] = df['رقم_بطاقة_المعالج'].astype(str).str.replace('.0', '', regex=False).str.strip().str.upper()
    df['rel_raw'] = df['الصلة'].astype(str).str.strip()

    # 1. Identify the TRUE principal for each family group (Employee Name)
    # The true principal ID is the one that has NO suffix (like H1, D1, S1 etc)
    # or the one where card == p_card in the original file.
    
    family_p_cards = {} # n2 -> principal_card
    for p_name, group in df.groupby('n2'):
        if not p_name: continue
        # Find the row that represents the principal (Employee)
        # Usually where card has no suffix or name == employee_name
        p_row = group[group['n1'] == p_name]
        if not p_row.empty:
            family_p_cards[p_name] = p_row.iloc[0]['card']
        else:
            # First row in group, take the base of its card
            base_card = group['card'].iloc[0].split('H')[0].split('D')[0].split('S')[0].split('M')[0]
            family_p_cards[p_name] = base_card

    # 2. Process Records with Corrected Linkage
    processed = []
    seen_names = set()
    
    for _, row in df.iterrows():
        n1 = row['n1']
        n2 = row['n2']
        card = row['card']
        rel = map_relationship(row['rel_raw'])
        p_card = family_p_cards.get(n2)
        
        is_p = (card == p_card) or (n1 == n2 and not rel)
        
        # Name Repair
        real_name = n1
        if not is_p and n1 == n2 and n1 != "":
            # If the member name matches the father name but it's a dependent,
            # this row is likely an 'Add missing info' row or has the name in another column.
            # However, looking at the data, we should try to skip absolute duplicates later.
            pass
            
        if not real_name: continue

        record = {
            'full_name': real_name,
            'employer': 'المنطقة الحرة جليانة' if is_p else '',
            'principal_card_number': p_card if not is_p else '',
            'relationship': rel,
            'card_number': card,
            'SORT_ORDER': 0 if is_p else 1,
            'SORT_GROUP': n2
        }
        
        # Deduplicate same name within the same family to avoid mistakes
        name_key = (record['full_name'].lower(), record['SORT_GROUP'].lower())
        if name_key not in seen_names:
            processed.append(record)
            seen_names.add(name_key)

    # 3. Add Virtual Principals if they exist in n2 but not as members in n1
    member_cards = {r['card_number'] for r in processed}
    for p_name, p_card in family_p_cards.items():
        if p_card not in member_cards and p_name.lower() not in {r['full_name'].lower() for r in processed}:
            processed.append({
                'full_name': p_name,
                'employer': 'المنطقة الحرة جليانة',
                'principal_card_number': '',
                'relationship': '',
                'card_number': p_card,
                'SORT_ORDER': 0,
                'SORT_GROUP': p_name
            })

    final_df = pd.DataFrame(processed)
    
    # User's strict requirement: 5400 rows, no duplicates.
    # Deduplicate strictly by name across the whole file
    final_df = final_df.sort_values(by=['full_name', 'employer'], ascending=[True, False])
    final_df = final_df.drop_duplicates(subset=['full_name'], keep='first')
    
    print(f"Final Row Count: {len(final_df)}")
    final_df = final_df.sort_values(by=['SORT_GROUP', 'SORT_ORDER'])

    # 4. Fill Template
    output_path = r'd:\tba_waad_system-main\tba_waad_system-main\جليانة.xlsx'
    template_path = r'd:\tba_waad_system-main\tba_waad_system-main\members_template (3).xlsx'
    wb = load_workbook(template_path)
    ws = wb['Data']
    
    # Headers
    column_mapping = {'full_name': 'الاسم الكامل', 'employer': 'جهة العمل', 'principal_card_number': 'رقم بطاقة الرئيسي', 'relationship': 'القرابة', 'card_number': 'رقم البطاقة'}
    for col_idx, header in enumerate(column_mapping.values(), 1):
        ws.cell(row=1, column=col_idx, value=header)
    for r in ws.iter_rows(min_row=2):
        for cell in r: cell.value = None

    row_idx = 3 
    for _, series in final_df.iterrows():
        ws.cell(row=row_idx, column=1, value=series['full_name'])
        ws.cell(row=row_idx, column=2, value=series['employer'])
        ws.cell(row=row_idx, column=3, value=series['principal_card_number'])
        ws.cell(row=row_idx, column=4, value=series['relationship'])
        ws.cell(row=row_idx, column=5, value=series['card_number'])
        row_idx += 1
        
    wb.save(output_path)
    print(f"SUCCESS: جليانة.xlsx generated with {len(final_df)} rows and UNIFIED linkage.")

except Exception as e:
    import traceback
    traceback.print_exc()
    sys.exit(1)
