
import pandas as pd
import sys
import os
from openpyxl import load_workbook

def normalize_name(name):
    if not isinstance(name, str): return ""
    name = str(name).strip()
    if name.lower() in ['الاسم', 'اسم الموظف', 'الرقم الوظيفي', 'nan', 'none', '']: 
        return ""
    # Remove extra spaces inside name
    return " ".join(name.split())

def map_relationship(rel):
    rel = str(rel).strip()
    if not rel or rel == 'nan' or rel == 'موظف' or rel == 'Principal' or rel == 'نفسه' or rel == 'None' or rel == 'الصلة':
        return ""
    
    mapping = {
        'ابن': 'SON',
        'ابنة': 'DAUGHTER',
        'بنت': 'DAUGHTER',
        'زوج': 'HUSBAND',
        'زوجة': 'WIFE',
        'زوجه': 'WIFE',
        'أب': 'FATHER',
        'اب': 'FATHER',
        'أم': 'MOTHER',
        'ام': 'MOTHER',
        'أخ': 'BROTHER',
        'اخ': 'BROTHER',
        'أخت': 'SISTER',
        'اخت': 'SISTER'
    }
    return mapping.get(rel, rel)

try:
    source_file = r'd:\tba_waad_system-main\tba_waad_system-main\القائمة_الموحدة_جليانة_جاهزة_للرفع_دقيقة_100.xlsx'
    df = pd.read_excel(source_file)
    print(f"Original source rows: {len(df)}")
    
    # 1. Standardize and Clean
    df['name_norm'] = df['الاسم'].apply(normalize_name)
    df['emp_name_norm'] = df['اسم_الموظف'].apply(normalize_name)
    df['card_clean'] = df['رقم_بطاقة_المعالج'].astype(str).str.replace('.0', '', regex=False).str.strip().str.replace('nan', '', case=False).str.upper()
    df['p_card_clean'] = df['رقم_البطاقة_التأمينية'].astype(str).str.replace('.0', '', regex=False).str.strip().str.replace('nan', '', case=False).str.upper()
    
    # 2. FIX NAME COLLISIONS
    # We need to ensure that each unique name (lower case) is associated with ONLY ONE Card Number
    # If a name has multiple cards, we'll pick the first one and remap all dependents to it.
    name_to_card_map = {}
    print("Detecting name collisions...")
    
    # Only consider rows where name is actually a principal name in the source
    # (In this file, 'اسم_الموظف' is the principal's name)
    for _, row in df.iterrows():
        name_key = row['emp_name_norm'].lower()
        card_val = row['p_card_clean']
        if name_key and card_val and len(card_val) > 3:
            if name_key not in name_to_card_map:
                name_to_card_map[name_key] = card_val
            # If already mapped, we keep the first one to ensure stability
            
    # Remap p_card_clean for all rows based on the name_to_card_map
    def remap_p_card(row):
        name_key = row['emp_name_norm'].lower()
        return name_to_card_map.get(name_key, row['p_card_clean'])

    df['p_card_clean'] = df.apply(remap_p_card, axis=1)
    
    # 3. Identify Families and Missing Principals
    unique_parent_cards = df['p_card_clean'][df['p_card_clean'].str.len() > 3].unique()
    existing_principals = df['card_clean'][df['card_clean'].isin(unique_parent_cards)].unique()
    missing_principals = set(unique_parent_cards) - set(existing_principals)
    
    parent_to_emp_name = {}
    for _, row in df.iterrows():
        p_card = row['p_card_clean']
        e_name = row['emp_name_norm']
        if p_card and e_name and p_card not in parent_to_emp_name:
            parent_to_emp_name[p_card] = e_name

    # 4. Build Final Data
    final_data = []
    
    # Add existing rows
    for _, row in df.iterrows():
        name = row['name_norm']
        if not name: continue
        
        card = row['card_clean']
        parent = row['p_card_clean']
        is_p = (card == parent) and (card != "")
        
        rel = map_relationship(row['الصلة'])
        
        final_data.append({
            'full_name': name,
            'employer': 'المنطقة الحرة جليانة' if is_p else '',
            'principal_card_number': parent if not is_p else '',
            'relationship': rel,
            'card_number': card,
            'SORT_ORDER': 0 if is_p else 1,
            'SORT_GROUP': parent if parent else card
        })

    # Add missing principals
    for m_card in missing_principals:
        p_name = parent_to_emp_name.get(m_card, f"موظف {m_card}")
        final_data.append({
            'full_name': p_name,
            'employer': 'المنطقة الحرة جليانة',
            'principal_card_number': '',
            'relationship': '',
            'card_number': m_card,
            'SORT_ORDER': 0,
            'SORT_GROUP': m_card
        })

    final_df = pd.DataFrame(final_data)
    
    # 5. DEDUPLICATE BY NAME AND EMPLOYER (Java style)
    # But we MUST keep the Principal row if a name appears as both P and D or multiple Ps
    # Sort: Principal rows (employer not empty) first
    final_df['is_p'] = final_df['employer'].apply(lambda x: 1 if x else 0)
    final_df = final_df.sort_values(by=['full_name', 'is_p'], ascending=[True, False])
    
    # Deduplicate by name (within same employer - here all are same employer or empty)
    # For safety, we deduplicate by (full_name) because Java uses that. 
    # But wait, if we have two different people with the same name, Java will skip one.
    # To fix this, we'll add a suffix if card numbers are different! 
    
    print("Deduplicating names and handling collisions...")
    unique_records = []
    seen_names = {} # name -> card_number
    
    # Sort back to family grouping first to keep family together
    final_df = final_df.sort_values(by=['SORT_GROUP', 'SORT_ORDER'], ascending=[True, True])
    
    for _, row in final_df.iterrows():
        name_key = row['full_name'].strip().lower()
        card = row['card_number']
        
        if name_key in seen_names:
            if seen_names[name_key] != card:
                # Name collision with different card! 
                # We modify the name slightly to allow Java to import it
                row['full_name'] = f"{row['full_name']} ({card})"
        else:
            seen_names[name_key] = card
            
        unique_records.append(row)

    final_df = pd.DataFrame(unique_records)
    
    # Final check: any row with empty card number should be removed
    final_df = final_df[final_df['card_number'].str.strip() != ""]
    
    # Deduplicate cards (absolute safety)
    final_df = final_df.drop_duplicates(subset=['card_number'])
    
    # Sort by family grouping
    final_df = final_df.sort_values(by=['SORT_GROUP', 'SORT_ORDER'], ascending=[True, True])

    print(f"Total Rows for output: {len(final_df)}")
    
    # 6. Fill Template
    output_path = r'd:\tba_waad_system-main\tba_waad_system-main\جليانة.xlsx'
    template_path = r'd:\tba_waad_system-main\tba_waad_system-main\members_template (3).xlsx'
    
    wb = load_workbook(template_path)
    ws = wb['Data']
    
    # Set headers
    column_mapping = {
        'full_name': 'الاسم الكامل',
        'employer': 'جهة العمل',
        'principal_card_number': 'رقم بطاقة الرئيسي',
        'relationship': 'القرابة',
        'card_number': 'رقم البطاقة'
    }
    headers = list(column_mapping.values())
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)

    # Clear
    for r in ws.iter_rows(min_row=2):
        for cell in r: cell.value = None

    # Fill
    row_idx = 3 # Data starts on Row 3
    for _, series in final_df.iterrows():
        ws.cell(row=row_idx, column=1, value=series['full_name'])
        ws.cell(row=row_idx, column=2, value=series['employer'])
        ws.cell(row=row_idx, column=3, value=series['principal_card_number'])
        ws.cell(row=row_idx, column=4, value=series['relationship'])
        ws.cell(row=row_idx, column=5, value=series['card_number'])
        row_idx += 1
        
    wb.save(output_path)
    print(f"SUCCESS: جليانة.xlsx V10 generated with {len(final_df)} rows.")

except Exception as e:
    import traceback
    traceback.print_exc()
    sys.exit(1)
