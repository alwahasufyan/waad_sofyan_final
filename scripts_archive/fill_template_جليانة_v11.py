
import pandas as pd
import sys
import os
from openpyxl import load_workbook

def normalize_name(name):
    if not isinstance(name, str): return ""
    name = str(name).strip()
    if name.lower() in ['الاسم', 'اسم الموظف', 'الرقم الوظيفي', 'nan', 'none', '']: 
        return ""
    return " ".join(name.split())

def map_relationship(rel):
    rel = str(rel).strip()
    if not rel or rel == 'nan' or rel == 'موظف' or rel == 'Principal' or rel == 'نفسه' or rel == 'None' or rel == 'الصلة':
        return ""
    
    mapping = {
        'ابن': 'SON', 'ابنة': 'DAUGHTER', 'بنت': 'DAUGHTER',
        'زوج': 'HUSBAND', 'زوجة': 'WIFE', 'زوجه': 'WIFE',
        'أب': 'FATHER', 'اب': 'FATHER', 'أم': 'MOTHER', 'ام': 'MOTHER',
        'أخ': 'BROTHER', 'اخ': 'BROTHER', 'أخت': 'SISTER', 'اخت': 'SISTER'
    }
    return mapping.get(rel, rel)

try:
    source_file = r'd:\tba_waad_system-main\tba_waad_system-main\القائمة_الموحدة_جليانة_جاهزة_للرفع_دقيقة_100.xlsx'
    df = pd.read_excel(source_file)
    print(f"Original source rows: {len(df)}")
    
    # 1. Standardize basics
    df['n1'] = df['الاسم'].apply(normalize_name)
    df['n2'] = df['اسم_الموظف'].apply(normalize_name)
    df['card'] = df['رقم_بطاقة_المعالج'].astype(str).str.replace('.0', '', regex=False).str.strip().str.upper()
    df['p_card'] = df['رقم_البطاقة_التأمينية'].astype(str).str.replace('.0', '', regex=False).str.strip().str.upper()
    df['rel_raw'] = df['الصلة'].astype(str).str.strip()
    
    # 2. Map Family Principals
    # A family is defined by p_card. Its principal name is the one where n1==n2 or rel_raw is 'موظف'
    family_principals = {}
    for p_card, group in df.groupby('p_card'):
        if not p_card or p_card == 'NAN': continue
        
        # Look for the principal row in this family
        p_row = group[(group['n1'] == group['n2']) | (group['rel_raw'].isin(['موظف', 'Principal', 'نفسه']))]
        if not p_row.empty:
            family_principals[p_card] = p_row.iloc[0]['n1']
        else:
            # Fallback: find the most common name among 'n2' (usually the principal name column in some sections)
            family_principals[p_card] = group['n1'].iloc[0]

    # 3. Extract Real Member Names
    processed_rows = []
    for _, row in df.iterrows():
        p_card = row['p_card']
        p_name = family_principals.get(p_card, "")
        
        # Decide which column is the member name
        # If n1 is the principal name, and n2 is different, n2 is the member.
        # Otherwise, n1 is the member.
        name = row['n1']
        if p_name and row['n1'] == p_name and row['n2'] != p_name and row['n2'] != "":
            # Special case for flipped data (like Mohammed Masoud's family)
            name = row['n2']
        elif not name and row['n2']:
            name = row['n2']
            
        if not name: continue
        
        rel = map_relationship(row['rel_raw'])
        is_p = (row['card'] == p_card) or (not rel)
        
        processed_rows.append({
            'full_name': name,
            'employer': 'المنطقة الحرة جليانة' if is_p else '',
            'principal_card_number': p_card if not is_p else '',
            'relationship': rel,
            'card_number': row['card'],
            'SORT_ORDER': 0 if is_p else 1,
            'SORT_GROUP': p_card
        })

    # 4. Handle Missing Principals
    existing_cards = {r['card_number'] for r in processed_rows}
    all_p_cards = {r['principal_card_number'] for r in processed_rows if r['principal_card_number']}
    missing_p_cards = all_p_cards - existing_cards
    
    for m_card in missing_p_cards:
        m_name = family_principals.get(m_card, f"موظف {m_card}")
        processed_rows.append({
            'full_name': m_name,
            'employer': 'المنطقة الحرة جليانة',
            'principal_card_number': '',
            'relationship': '',
            'card_number': m_card,
            'SORT_ORDER': 0,
            'SORT_GROUP': m_card
        })

    # 5. STRICT DEDUPLICATION BY NAME
    # The user is annoyed by repetition. 
    final_df = pd.DataFrame(processed_rows)
    
    # Sort to prefer Principal rows and non-empty relationships
    final_df['has_rel'] = final_df['relationship'].apply(lambda x: 1 if x else 0)
    final_df['is_p'] = final_df['employer'].apply(lambda x: 1 if x else 0)
    
    # Sort: Principals first, then rows with relationships
    final_df = final_df.sort_values(by=['full_name', 'is_p', 'has_rel'], ascending=[True, False, False])
    
    # Deduplicate by Name
    print(f"Rows before name-deduplication: {len(final_df)}")
    final_df = final_df.drop_duplicates(subset=['full_name'], keep='first')
    print(f"Rows after name-deduplication: {len(final_df)}")
    
    # 6. Final cleanup of Relationships
    # If a row has principal_card_number but NO relationship, set a default
    def fix_rel(row):
        if row['principal_card_number'] and not row['relationship']:
            return 'DEPENDENT'
        return row['relationship']
    final_df['relationship'] = final_df.apply(fix_rel, axis=1)

    # Sort by family grouping
    final_df = final_df.sort_values(by=['SORT_GROUP', 'SORT_ORDER'], ascending=[True, True])

    # 7. Fill Template
    output_path = r'd:\tba_waad_system-main\tba_waad_system-main\جليانة.xlsx'
    template_path = r'd:\tba_waad_system-main\tba_waad_system-main\members_template (3).xlsx'
    
    wb = load_workbook(template_path)
    ws = wb['Data']
    
    # Clear and set headers
    column_mapping = {'full_name': 'الاسم الكامل', 'employer': 'جهة العمل', 'principal_card_number': 'رقم بطاقة الرئيسي', 'relationship': 'القرابة', 'card_number': 'رقم البطاقة'}
    for col_idx, header in enumerate(column_mapping.values(), 1):
        ws.cell(row=1, column=col_idx, value=header)
    for r in ws.iter_rows(min_row=2):
        for cell in r: cell.value = None

    # Fill
    row_idx = 3 # Start on row 3
    for _, series in final_df.iterrows():
        ws.cell(row=row_idx, column=1, value=series['full_name'])
        ws.cell(row=row_idx, column=2, value=series['employer'])
        ws.cell(row=row_idx, column=3, value=series['principal_card_number'])
        ws.cell(row=row_idx, column=4, value=series['relationship'])
        ws.cell(row=row_idx, column=5, value=series['card_number'])
        row_idx += 1
        
    wb.save(output_path)
    print(f"SUCCESS: جليانة.xlsx V11 generated with {len(final_df)} UNIQUE names.")

except Exception as e:
    import traceback
    traceback.print_exc()
    sys.exit(1)
