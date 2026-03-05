
import pandas as pd
import sys
import os
from openpyxl import load_workbook

def normalize_name(name):
    if not isinstance(name, str): return ""
    name = str(name).strip()
    if name.lower() in ['الاسم', 'اسم الموظف', 'الرقم الوظيفي', 'nan', 'none', '']: 
        return ""
    return " ".join(name.split())

def map_relationship(rel):
    rel = str(rel).strip()
    if not rel or rel == 'nan' or rel == 'موظف' or rel == 'Principal' or rel == 'نفسه' or rel == 'None' or rel == 'الصلة':
        return ""
    
    mapping = {
        'ابن': 'SON', 'ابنة': 'DAUGHTER', 'بنت': 'DAUGHTER',
        'زوج': 'HUSBAND', 'زوجة': 'WIFE', 'زوجه': 'WIFE',
        'أب': 'FATHER', 'اب': 'FATHER', 'أم': 'MOTHER', 'ام': 'MOTHER',
        'أخ': 'BROTHER', 'اخ': 'BROTHER', 'أخت': 'SISTER', 'اخت': 'SISTER'
    }
    return mapping.get(rel, rel)

try:
    source_file = r'd:\tba_waad_system-main\tba_waad_system-main\القائمة_الموحدة_جليانة_جاهزة_للرفع_دقيقة_100.xlsx'
    df = pd.read_excel(source_file)
    print(f"Original source rows: {len(df)}")
    
    # Standardize basic columns
    df['n1'] = df['الاسم'].apply(normalize_name)
    df['n2'] = df['اسم_الموظف'].apply(normalize_name)
    df['card'] = df['رقم_بطاقة_المعالج'].astype(str).str.replace('.0', '', regex=False).str.strip().str.upper()
    df['p_card'] = df['رقم_البطاقة_التأمينية'].astype(str).str.replace('.0', '', regex=False).str.strip().str.upper()
    df['rel_raw'] = df['الصلة'].astype(str).str.strip()
    
    # 1. Map Family Principals correctly
    family_p_names = {}
    for p_card, group in df.groupby('p_card'):
        if not p_card or p_card == 'NAN': continue
        # Principal is the one where card == p_card
        p_row = group[group['card'] == p_card]
        if not p_row.empty:
            family_p_names[p_card] = p_row.iloc[0]['n1']
        else:
            # Fallback to the common principal name in this group
            family_p_names[p_card] = group['n1'].iloc[0]

    # 2. Extract and Repair Names
    final_rows = []
    for _, row in df.iterrows():
        card = row['card']
        p_card = row['p_card']
        n1 = row['n1']
        n2 = row['n2']
        p_name = family_p_names.get(p_card, "")
        
        is_p = (card == p_card) or (card != "" and p_card == "")
        rel = map_relationship(row['rel_raw'])
        
        # Name Repair Logic
        if is_p:
            real_name = n1
        else:
            # If the name column (n1) matches the father's name, then the real name is in n2 (Dependent's name)
            if n1 == p_name and n2 != p_name and n2 != "":
                real_name = n2
            else:
                real_name = n1
                
        if not real_name: continue
        
        final_rows.append({
            'full_name': real_name,
            'employer': 'المنطقة الحرة جليانة' if is_p else '',
            'principal_card_number': p_card if not is_p else '',
            'relationship': rel,
            'card_number': card,
            'SORT_ORDER': 0 if is_p else 1,
            'SORT_GROUP': p_card if p_card else card
        })

    # 3. Add Missing Principals if any family has no principal row
    existing_cards = {r['card_number'] for r in final_rows}
    needed_p_cards = {r['principal_card_number'] for r in final_rows if r['principal_card_number']}
    missing_p_cards = needed_p_cards - existing_cards
    
    for m_card in missing_p_cards:
        m_name = family_p_names.get(m_card, f"موظف {m_card}")
        final_rows.append({
            'full_name': m_name,
            'employer': 'المنطقة الحرة جليانة',
            'principal_card_number': '',
            'relationship': '',
            'card_number': m_card,
            'SORT_ORDER': 0,
            'SORT_GROUP': m_card
        })

    # 4. Handle Name Conflicts (Duplicates)
    # If two different people have the same name, we must make them unique for Java
    final_df = pd.DataFrame(final_rows)
    # Deduplicate absolute duplicates (same name, same card, same parent)
    final_df = final_df.drop_duplicates(subset=['full_name', 'card_number', 'principal_card_number'])
    
    print("Making unique names for remaining collisions...")
    unique_final = []
    seen_names = {} # name -> card
    
    # Sort to process principals first
    final_df['is_p'] = final_df['employer'].apply(lambda x: 1 if x else 0)
    final_df = final_df.sort_values(by=['is_p', 'full_name'], ascending=[False, True])
    
    for _, row in final_df.iterrows():
        name_key = row['full_name'].lower().strip()
        card = row['card_number']
        
        if name_key in seen_names:
            if seen_names[name_key] != card:
                # Same name, different person (card)
                # We add a hidden-ish suffix or just the card number if it's a real collision
                # To keep it clean, we only do this if it's NOT the same record
                row['full_name'] = f"{row['full_name']} *{card[-4:]}" # Small identification suffix
        else:
            seen_names[name_key] = card
        unique_final.append(row)
        
    final_df = pd.DataFrame(unique_final)
    
    # 5. Final Touch: Ensure Relationship for all dependents
    def fix_missing_rel(row):
        if row['principal_card_number'] and not row['relationship']:
            return 'DEPENDENT'
        return row['relationship']
    final_df['relationship'] = final_df.apply(fix_missing_rel, axis=1)

    # Sort for grouping families visually
    final_df = final_df.sort_values(by=['SORT_GROUP', 'SORT_ORDER'], ascending=[True, True])
    print(f"Total rows for output: {len(final_df)}")

    # 6. Fill Template
    output_path = r'd:\tba_waad_system-main\tba_waad_system-main\جليانة.xlsx'
    template_path = r'd:\tba_waad_system-main\tba_waad_system-main\members_template (3).xlsx'
    
    wb = load_workbook(template_path)
    ws = wb['Data']
    
    # Headers
    column_mapping = {'full_name': 'الاسم الكامل', 'employer': 'جهة العمل', 'principal_card_number': 'رقم بطاقة الرئيسي', 'relationship': 'القرابة', 'card_number': 'رقم البطاقة'}
    for col_idx, header in enumerate(column_mapping.values(), 1):
        ws.cell(row=1, column=col_idx, value=header)
    # Clear
    for r in ws.iter_rows(min_row=2):
        for cell in r: cell.value = None

    # Fill
    row_idx = 3 
    for _, series in final_df.iterrows():
        ws.cell(row=row_idx, column=1, value=series['full_name'])
        ws.cell(row=row_idx, column=2, value=series['employer'])
        ws.cell(row=row_idx, column=3, value=series['principal_card_number'])
        ws.cell(row=row_idx, column=4, value=series['relationship'])
        ws.cell(row=row_idx, column=5, value=series['card_number'])
        row_idx += 1
        
    wb.save(output_path)
    print(f"SUCCESS: جليانة.xlsx V12 generated with {len(final_df)} rows and repaired names.")

except Exception as e:
    import traceback
    traceback.print_exc()
    sys.exit(1)
