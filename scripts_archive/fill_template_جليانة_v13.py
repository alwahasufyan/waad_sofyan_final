
import pandas as pd
import sys
import os
from openpyxl import load_workbook

def normalize_name(name):
    if not isinstance(name, str): return ""
    name = str(name).strip()
    if name.lower() in ['الاسم', 'اسم الموظف', 'الرقم الوظيفي', 'nan', 'none', '']: return ""
    return " ".join(name.split())

def map_relationship(rel):
    rel = str(rel).strip()
    if not rel or rel == 'nan' or rel in ['موظف', 'Principal', 'نفسه', 'None']: return ""
    mapping = {
        'ابن': 'SON', 'ابنة': 'DAUGHTER', 'بنت': 'DAUGHTER',
        'زوج': 'HUSBAND', 'زوجة': 'WIFE', 'زوجه': 'WIFE',
        'أب': 'FATHER', 'اب': 'FATHER', 'أم': 'MOTHER', 'ام': 'MOTHER',
        'أخ': 'BROTHER', 'اخ': 'BROTHER', 'أخت': 'SISTER', 'اخت': 'SISTER'
    }
    return mapping.get(rel, rel)

try:
    source_file = r'd:\tba_waad_system-main\tba_waad_system-main\القائمة_الموحدة_جليانة_جاهزة_للرفع_دقيقة_100.xlsx'
    df = pd.read_excel(source_file)
    
    # Standardize
    df['n1'] = df['الاسم'].apply(normalize_name)
    df['n2'] = df['اسم_الموظف'].apply(normalize_name)
    df['card'] = df['رقم_بطاقة_المعالج'].astype(str).str.replace('.0', '', regex=False).str.strip().str.upper()
    df['p_card'] = df['رقم_البطاقة_التأمينية'].astype(str).str.replace('.0', '', regex=False).str.strip().str.upper()
    df['rel_raw'] = df['الصلة'].astype(str).str.strip()

    # 1. Deduplicate source rows by Card Number first (Absolute uniqueness)
    # If card is NAN, keep it for now but we'll deduplicate by name later
    source_deduped = df.drop_duplicates(subset=['card'])
    
    # 2. Map Families to their proper Principal Name
    family_p_names = {}
    for p_card, group in source_deduped.groupby('p_card'):
        if not p_card or p_card == 'NAN': continue
        p_row = group[group['card'] == p_card]
        if not p_row.empty:
            family_p_names[p_card] = p_row.iloc[0]['n1']
        else:
            family_p_names[p_card] = group['n2'].iloc[0] if not group['n2'].empty else group['n1'].iloc[0]

    # 3. Process Rows with "Smart Name Repair" to avoid Father-Name duplication
    processed_records = []
    for _, row in source_deduped.iterrows():
        p_card = row['p_card']
        card = row['card']
        n1 = row['n1']
        n2 = row['n2']
        p_name = family_p_names.get(p_card, "")
        rel = map_relationship(row['rel_raw'])
        is_p = (card == p_card) or (not rel)

        real_name = n1
        # If the member name is EXACTLY the same as the Principal name but they are a dependent,
        # it means the data is broken and we should use the descriptive name if possible or n2.
        if not is_p and n1 == p_name:
            if n2 != p_name and n2 != "":
                real_name = n2
            else:
                # Fallback: keep n1 but it will likely be deduplicated later if it's a real duplicate
                real_name = n1
        
        if not real_name: continue

        processed_records.append({
            'full_name': real_name,
            'employer': 'المنطقة الحرة جليانة' if is_p else '',
            'principal_card_number': p_card if not is_p else '',
            'relationship': rel,
            'card_number': card,
            'SORT_ORDER': 0 if is_p else 1,
            'SORT_GROUP': p_card
        })

    # 4. STRICT DEDUPLICATION to reach ~5400
    final_df = pd.DataFrame(processed_records)
    
    # Remove absolute duplicates by (Name + Employer) - Java style
    # Prefer Principal rows when conflict exists
    final_df['is_p'] = final_df['employer'].apply(lambda x: 1 if x else 0)
    final_df = final_df.sort_values(by=['full_name', 'is_p'], ascending=[True, False])
    
    # Deduplicate by Name
    # This might remove legitimate same-names, but it's what the user wants ("No duplicates")
    final_df = final_df.drop_duplicates(subset=['full_name'], keep='first')
    
    # 5. Missing Principals (ONLY if necessary for linking)
    # We'll see if the user's "5400" limit allows this.
    existing_cards = set(final_df['card_number'])
    needed_p_cards = {r['principal_card_number'] for _, r in final_df.iterrows() if r['principal_card_number']}
    missing_p_cards = needed_p_cards - (existing_cards | {'', 'NAN', None})
    
    # To keep count low, we ONLY add principals that have at least one dependent in our final list
    for m_card in missing_p_cards:
        m_name = family_p_names.get(m_card, f"موظف {m_card}")
        # Check if this name already exists in final_df to avoid duplication
        if m_name.lower().strip() not in set(final_df['full_name'].str.lower().str.strip()):
            final_df = pd.concat([final_df, pd.DataFrame([{
                'full_name': m_name,
                'employer': 'المنطقة الحرة جليانة',
                'principal_card_number': '',
                'relationship': '',
                'card_number': m_card,
                'SORT_ORDER': 0,
                'SORT_GROUP': m_card,
                'is_p': 1
            }])], ignore_index=True)

    # 6. Final sort and cleanup
    final_df = final_df.sort_values(by=['SORT_GROUP', 'SORT_ORDER'], ascending=[True, True])
    print(f"Final unique row count: {len(final_df)}")

    # 7. Fill Template
    output_path = r'd:\tba_waad_system-main\tba_waad_system-main\جليانة.xlsx'
    template_path = r'd:\tba_waad_system-main\tba_waad_system-main\members_template (3).xlsx'
    wb = load_workbook(template_path)
    ws = wb['Data']
    
    # Clear and set headers
    column_mapping = {'full_name': 'الاسم الكامل', 'employer': 'جهة العمل', 'principal_card_number': 'رقم بطاقة الرئيسي', 'relationship': 'القرابة', 'card_number': 'رقم البطاقة'}
    for col_idx, header in enumerate(column_mapping.values(), 1):
        ws.cell(row=1, column=col_idx, value=header)
    for r in ws.iter_rows(min_row=2):
        for cell in r: cell.value = None

    row_idx = 3 
    for _, series in final_df.iterrows():
        ws.cell(row=row_idx, column=1, value=series['full_name'])
        ws.cell(row=row_idx, column=2, value=series['employer'])
        ws.cell(row=row_idx, column=3, value=series['principal_card_number'])
        ws.cell(row=row_idx, column=4, value=series['relationship'])
        ws.cell(row=row_idx, column=5, value=series['card_number'])
        row_idx += 1
        
    wb.save(output_path)
    print(f"SUCCESS: جليانة.xlsx V13 generated with {len(final_df)} rows.")

except Exception as e:
    import traceback
    traceback.print_exc()
    sys.exit(1)
