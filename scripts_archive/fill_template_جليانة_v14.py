
import pandas as pd
import sys
import os
from openpyxl import load_workbook

def normalize_name(name):
    if not isinstance(name, str): return ""
    name = str(name).strip()
    if name.lower() in ['الاسم', 'اسم الموظف', 'الرقم الوظيفي', 'nan', 'none', '']: return ""
    return " ".join(name.split())

def map_relationship(rel):
    rel = str(rel).strip()
    if not rel or rel == 'nan' or rel in ['موظف', 'Principal', 'نفسه', 'None']: return ""
    mapping = {
        'ابن': 'SON', 'ابنة': 'DAUGHTER', 'بنت': 'DAUGHTER',
        'زوج': 'HUSBAND', 'زوجة': 'WIFE', 'زوجه': 'WIFE',
        'أب': 'FATHER', 'اب': 'FATHER', 'أم': 'MOTHER', 'ام': 'MOTHER',
        'أخ': 'BROTHER', 'اخ': 'BROTHER', 'أخت': 'SISTER', 'اخت': 'SISTER'
    }
    return mapping.get(rel, rel)

try:
    source_file = r'd:\tba_waad_system-main\tba_waad_system-main\القائمة_الموحدة_جليانة_جاهزة_للرفع_دقيقة_100.xlsx'
    df = pd.read_excel(source_file)
    print(f"Original source rows: {len(df)}")
    
    # Standardize
    df['n1'] = df['الاسم'].apply(normalize_name)
    df['n2'] = df['اسم_الموظف'].apply(normalize_name)
    df['card'] = df['رقم_بطاقة_المعالج'].astype(str).str.replace('.0', '', regex=False).str.strip().str.upper()
    df['p_card'] = df['رقم_البطاقة_التأمينية'].astype(str).str.replace('.0', '', regex=False).str.strip().str.upper()
    df['rel_raw'] = df['الصلة'].astype(str).str.strip()

    # 1. Identify Principal Name for each Family
    # A family is p_card. Its principal name is the one that appears most often in either n1 or n2 
    # OR specifically the one marked as 'موظف'
    family_p_names = {}
    for p_card, group in df.groupby('p_card'):
        if not p_card or p_card == 'NAN': continue
        
        # Strategy: find name where card == p_card
        p_row = group[group['card'] == p_card]
        if not p_row.empty:
            family_p_names[p_card] = p_row.iloc[0]['n1']
        else:
            # Fallback: Find which name appears in almost every 'p_card' row in n1 (as parent name)
            # For dependents, n1 is usually 'Parent' and n2 is 'Member' (or vice versa)
            # Let's see which name is common across the whole subgroup in either column
            names = list(group['n1']) + list(group['n2'])
            names = [n for n in names if n]
            if names:
                # Most frequent name is likely the principal's name
                family_p_names[p_card] = max(set(names), key=names.count)

    # 2. Extract Real Members
    processed_records = []
    seen_cards = set()

    for _, row in df.iterrows():
        p_card = row['p_card']
        card = row['card']
        if card in seen_cards and card != 'NAN': continue
        
        n1 = row['n1']
        n2 = row['n2']
        p_name = family_p_names.get(p_card, "")
        rel = map_relationship(row['rel_raw'])
        is_p = (card == p_card) or (not rel)

        # Logic: Pick the name that is NOT the principal name, 
        # unless it's a principal row or both columns have the same name.
        real_name = n1
        if not is_p:
            if n1 == p_name and n2 != "" and n2 != p_name:
                real_name = n2
            elif n2 == p_name and n1 != "" and n1 != p_name:
                real_name = n1
            # If both are p_name but it's a dependent, we have a problem but we keep n1
        
        if not real_name: continue

        processed_records.append({
            'full_name': real_name,
            'employer': 'المنطقة الحرة جليانة' if is_p else '',
            'principal_card_number': p_card if not is_p else '',
            'relationship': rel,
            'card_number': card,
            'SORT_ORDER': 0 if is_p else 1,
            'SORT_GROUP': p_card if p_card else card
        })
        if card != 'NAN': seen_cards.add(card)

    # 3. Add Missing Principals
    # If a p_card is referenced but no row has card == p_card
    all_member_cards = {r['card_number'] for r in processed_records}
    all_parent_cards = {r['principal_card_number'] for r in processed_records if r['principal_card_number']}
    missing_p_cards = all_parent_cards - all_member_cards
    
    for m_card in missing_p_cards:
        if m_card in ['NAN', '', None]: continue
        m_name = family_p_names.get(m_card, f"موظف {m_card}")
        processed_records.append({
            'full_name': m_name,
            'employer': 'المنطقة الحرة جليانة',
            'principal_card_number': '',
            'relationship': '',
            'card_number': m_card,
            'SORT_ORDER': 0,
            'SORT_GROUP': m_card
        })

    # 4. Final Cleanup: Deduplicate by Name+Card to be super safe
    final_df = pd.DataFrame(processed_records)
    print(f"Total rows before final dedup: {len(final_df)}")
    
    # Deduplicate by full name (Java style) to ensure no 'unimported' rows
    final_df = final_df.drop_duplicates(subset=['full_name'], keep='first')
    
    print(f"Final unique row count: {len(final_df)}")

    # 5. Fill Template
    output_path = r'd:\tba_waad_system-main\tba_waad_system-main\جليانة.xlsx'
    template_path = r'd:\tba_waad_system-main\tba_waad_system-main\members_template (3).xlsx'
    wb = load_workbook(template_path)
    ws = wb['Data']
    
    column_mapping = {'full_name': 'الاسم الكامل', 'employer': 'جهة العمل', 'principal_card_number': 'رقم بطاقة الرئيسي', 'relationship': 'القرابة', 'card_number': 'رقم البطاقة'}
    for col_idx, header in enumerate(column_mapping.values(), 1):
        ws.cell(row=1, column=col_idx, value=header)
    for r in ws.iter_rows(min_row=2):
        for cell in r: cell.value = None

    row_idx = 3 
    # Group by family
    final_df = final_df.sort_values(by=['SORT_GROUP', 'SORT_ORDER'])
    for _, series in final_df.iterrows():
        ws.cell(row=row_idx, column=1, value=series['full_name'])
        ws.cell(row=row_idx, column=2, value=series['employer'])
        ws.cell(row=row_idx, column=3, value=series['principal_card_number'])
        ws.cell(row=row_idx, column=4, value=series['relationship'])
        ws.cell(row=row_idx, column=5, value=series['card_number'])
        row_idx += 1
        
    wb.save(output_path)
    print(f"SUCCESS: جليانة.xlsx V14 generated with {len(final_df)} rows.")

except Exception as e:
    import traceback
    traceback.print_exc()
    sys.exit(1)
