
import pandas as pd
import sys
import os
from openpyxl import load_workbook

def normalize_name(name):
    if not isinstance(name, str): return ""
    name = str(name).strip()
    if name.lower() in ['الاسم', 'اسم الموظف', 'الرقم الوظيفي', 'nan', 'none', '']: return ""
    return " ".join(name.split())

def map_relationship(rel):
    rel = str(rel).strip()
    if not rel or rel == 'nan' or rel in ['موظف', 'Principal', 'نفسه', 'None']: return ""
    mapping = {
        'ابن': 'SON', 'ابنة': 'DAUGHTER', 'بنت': 'DAUGHTER',
        'زوج': 'HUSBAND', 'زوجة': 'WIFE', 'زوجه': 'WIFE',
        'أب': 'FATHER', 'اب': 'FATHER', 'أم': 'MOTHER', 'ام': 'MOTHER',
        'أخ': 'BROTHER', 'اخ': 'BROTHER', 'أخت': 'SISTER', 'اخت': 'SISTER'
    }
    return mapping.get(rel, rel)

try:
    source_file = r'd:\tba_waad_system-main\tba_waad_system-main\القائمة_الموحدة_جليانة_جاهزة_للرفع_دقيقة_100.xlsx'
    df = pd.read_excel(source_file)
    print(f"Original source rows: {len(df)}")
    
    # Standardize
    df['n1'] = df['الاسم'].apply(normalize_name)
    df['n2'] = df['اسم_الموظف'].apply(normalize_name)
    df['card'] = df['رقم_بطاقة_المعالج'].astype(str).str.replace('.0', '', regex=False).str.strip().str.upper()
    
    # 1. GROUP BY EMPLOYEE NAME (n2) to unify families
    # The real anchor for a family is the employee name.
    family_groups = df.groupby('n2')
    
    processed_records = []
    seen_member_keys = set() # (name, card) to prevent duplicates

    for emp_name, group in family_groups:
        if not emp_name: continue
        
        # Determine the Principal for this family
        # The principal is ideally the row where n1 == n2 (Employee is the Member)
        p_rows = group[group['n1'] == emp_name]
        
        if not p_rows.empty:
            p_row = p_rows.iloc[0]
            p_card = p_row['card']
        else:
            # If no principal row exists, we MUST create/designate one
            # We'll use the card from the first person in group as base for a virtual principal if needed,
            # but usually one of them has a card that looks like a principal card (no suffix).
            # Let's pick the first card as principal card base.
            p_card = group['card'].iloc[0].split('H')[0].split('D')[0].split('S')[0].split('M')[0]
            # Add a virtual principal row
            processed_records.append({
                'full_name': emp_name,
                'employer': 'المنطقة الحرة جليانة',
                'principal_card_number': '',
                'relationship': '',
                'card_number': p_card,
                'SORT_ORDER': 0,
                'SORT_GROUP': emp_name
            })
            seen_member_keys.add((emp_name.lower(), p_card))

        # Now process all members in this group as dependents OF THIS p_card
        for _, row in group.iterrows():
            name = row['n1']
            card = row['card']
            rel = map_relationship(row['الصلة'])
            
            # If this is the principal row, we already handled it or will handle it
            is_actual_p = (name == emp_name)
            
            if is_actual_p:
                record = {
                    'full_name': name,
                    'employer': 'المنطقة الحرة جليانة',
                    'principal_card_number': '',
                    'relationship': '',
                    'card_number': card,
                    'SORT_ORDER': 0,
                    'SORT_GROUP': emp_name
                }
            else:
                # Force relationship to DEPENDENT if missing
                if not rel: rel = 'DEPENDENT'
                record = {
                    'full_name': name,
                    'employer': '',
                    'principal_card_number': p_card,
                    'relationship': rel,
                    'card_number': card,
                    'SORT_ORDER': 1,
                    'SORT_GROUP': emp_name
                }
                
            key = (record['full_name'].lower(), record['card_number'])
            if key not in seen_member_keys:
                processed_records.append(record)
                seen_member_keys.add(key)

    # Convert to DataFrame
    final_df = pd.DataFrame(processed_records)
    
    # Deduplicate strictly by name for the user's "Clean" requirement
    print(f"Total rows before name dedup: {len(final_df)}")
    final_df = final_df.drop_duplicates(subset=['full_name'], keep='first')
    print(f"Final clean row count: {len(final_df)}")

    # Sort
    final_df = final_df.sort_values(by=['SORT_GROUP', 'SORT_ORDER'])

    # 4. Fill Template
    output_path = r'd:\tba_waad_system-main\tba_waad_system-main\جليانة.xlsx'
    template_path = r'd:\tba_waad_system-main\tba_waad_system-main\members_template (3).xlsx'
    wb = load_workbook(template_path)
    ws = wb['Data']
    
    # Clear and set headers
    column_mapping = {'full_name': 'الاسم الكامل', 'employer': 'جهة العمل', 'principal_card_number': 'رقم بطاقة الرئيسي', 'relationship': 'القرابة', 'card_number': 'رقم البطاقة'}
    for col_idx, header in enumerate(column_mapping.values(), 1):
        ws.cell(row=1, column=col_idx, value=header)
    for r in ws.iter_rows(min_row=2):
        for cell in r: cell.value = None

    row_idx = 3 
    for _, series in final_df.iterrows():
        ws.cell(row=row_idx, column=1, value=series['full_name'])
        ws.cell(row=row_idx, column=2, value=series['employer'])
        ws.cell(row=row_idx, column=3, value=series['principal_card_number'])
        ws.cell(row=row_idx, column=4, value=series['relationship'])
        ws.cell(row=row_idx, column=5, value=series['card_number'])
        row_idx += 1
        
    wb.save(output_path)
    print(f"SUCCESS: جليانة.xlsx V15 generated with {len(final_df)} rows.")

except Exception as e:
    import traceback
    traceback.print_exc()
    sys.exit(1)
