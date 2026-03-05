import pandas as pd
import sys
import os
from openpyxl import load_workbook

def normalize_name(name):
    if not isinstance(name, str): return ""
    name = str(name)
    if name.lower() in ['الاسم', 'اسم الموظف', 'الرقم الوظيفي', 'nan', 'none']: 
        return ""
    return " ".join(name.strip().split())

def map_relationship(rel):
    rel = str(rel).strip()
    if rel in ['موظف', 'nan', 'Principal', 'نفسه'] or not rel or rel == 'None' or rel == 'الصلة':
        return ""
    
    mapping = {
        'ابن': 'SON',
        'ابنة': 'DAUGHTER',
        'بنت': 'DAUGHTER',
        'زوج': 'HUSBAND',
        'زوجة': 'WIFE',
        'زوجه': 'WIFE',
        'أب': 'FATHER',
        'اب': 'FATHER',
        'أم': 'MOTHER',
        'ام': 'MOTHER',
        'أخ': 'BROTHER',
        'اخ': 'BROTHER',
        'أخت': 'SISTER',
        'اخت': 'SISTER'
    }
    return mapping.get(rel, rel)

try:
    # 1. Process source data
    source_file = r'd:\tba_waad_system-main\tba_waad_system-main\القائمة_الموحدة_جليانة_جاهزة_للرفع_دقيقة_100.xlsx'
    df = pd.read_excel(source_file)
    
    df = df.dropna(how='all')
    df['الاسم'] = df['الاسم'].apply(normalize_name)
    df = df[df['الاسم'] != '']
    
    df['اسم_موظف_نرم':] = df['اسم_الموظف'].apply(normalize_name)
    df['الرقم_الوطني'] = df['الرقم_الوطني'].fillna('').astype(str).str.replace('.0', '', regex=False).str.strip()
    
    # Standardize card numbers
    # رقم_البطاقة_التأمينية: Principal Card Number
    # رقم_بطاقة_المعالج: Unique Card Number
    df['principal_card_key'] = df['رقم_البطاقة_التأمينية'].astype(str).str.replace('.0', '', regex=False).str.strip()
    df['unique_card_key'] = df['رقم_بطاقة_المعالج'].astype(str).str.replace('.0', '', regex=False).str.strip()
    
    dependent_rels = ['ابن', 'ابنة', 'بنت', 'زوج', 'زوجة', 'زوجه', 'أب', 'اب', 'أم', 'ام', 'أخ', 'اخ', 'أخت', 'اخت']
    df['is_principal'] = df.apply(lambda row: str(row['الصلة']).strip() not in dependent_rels, axis=1)
    
    # 2. Prepare data for template
    final_data = []
    for _, row in df.iterrows():
        final_data.append({
            'full_name': row['الاسم'],
            'employer': 'المنطقة الحرة جليانة' if row['is_principal'] else '',
            'principal_card_number': row['principal_card_key'] if not row['is_principal'] else '',
            'relationship': map_relationship(row['الصلة']),
            'card_number': row['unique_card_key'],
            'SORT_EMP': row['اسم_موظف_نرم':] if 'اسم_موظف_نرم': in df.columns else row['اسم_الموظف'],
            'SORT_TYPE': 1 if row['is_principal'] else 0
        })
    
    final_df = pd.DataFrame(final_data)
    # Deduplicate by unique_card_key (since that's the primary key in theory)
    final_df = final_df.drop_duplicates(subset=['card_number'])
    
    # Sort to ensure principals come before their dependents if possible
    final_df = final_df.sort_values(by=['SORT_EMP', 'SORT_TYPE'], ascending=[True, False])
    
    # 3. Load Template and Fill
    template_path = r'd:\tba_waad_system-main\tba_waad_system-main\members_template (3).xlsx'
    output_path = r'd:\tba_waad_system-main\tba_waad_system-main\جليانة.xlsx'
    
    wb = load_workbook(template_path)
    ws = wb['Data']
    
    # Clean sheet first (from row 2 down)
    if ws.max_row >= 2:
        for r in range(2, ws.max_row + 1):
            for c in range(1, 10): # Clear first 10 columns
                ws.cell(row=r, column=c).value = None

    row_idx = 2
    for _, series in final_df.iterrows():
        ws.cell(row=row_idx, column=1, value=series['full_name'])
        ws.cell(row=row_idx, column=2, value=series['employer'])
        ws.cell(row=row_idx, column=3, value=series['principal_card_number'])
        ws.cell(row=row_idx, column=4, value=series['relationship'])
        ws.cell(row=row_idx, column=5, value=series['card_number']) # This is the new column I supported in backend
        row_idx += 1
        
    wb.save(output_path)
    print(f"Data filled successfully. Final rows: {len(final_df)}")
    print(f"Saved to: {output_path}")

except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)
