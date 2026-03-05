import pandas as pd
import sys
import os
from openpyxl import load_workbook

def normalize_name(name):
    if not isinstance(name, str): return ""
    name = str(name)
    if name.lower() in ['الاسم', 'اسم الموظف', 'الرقم الوظيفي', 'nan', 'none']: 
        return ""
    return " ".join(name.strip().split())

def map_relationship(rel):
    rel = str(rel).strip()
    if rel in ['موظف', 'nan', 'Principal', 'نفسه'] or not rel or rel == 'None' or rel == 'الصلة':
        return ""
    
    mapping = {
        'ابن': 'SON',
        'ابنة': 'DAUGHTER',
        'بنت': 'DAUGHTER',
        'زوج': 'HUSBAND',
        'زوجة': 'WIFE',
        'زوجه': 'WIFE',
        'أب': 'FATHER',
        'اب': 'FATHER',
        'أم': 'MOTHER',
        'ام': 'MOTHER',
        'أخ': 'BROTHER',
        'اخ': 'BROTHER',
        'أخت': 'SISTER',
        'اخت': 'SISTER'
    }
    return mapping.get(rel, rel)

try:
    # 1. Process source data
    source_file = r'd:\tba_waad_system-main\tba_waad_system-main\القائمة_الموحدة_جليانة_جاهزة_للرفع_دقيقة_100.xlsx'
    df = pd.read_excel(source_file)
    
    df = df.dropna(how='all')
    df['الاسم'] = df['الاسم'].apply(normalize_name)
    df = df[df['الاسم'] != '']
    
    df['emp_name_norm'] = df['اسم_الموظف'].apply(normalize_name)
    
    # Standardize card numbers
    df['principal_card_key'] = df['رقم_البطاقة_التأمينية'].astype(str).str.replace('.0', '', regex=False).str.strip()
    df['unique_card_key'] = df['رقم_بطاقة_المعالج'].astype(str).str.replace('.0', '', regex=False).str.strip()
    
    df['principal_card_key'] = df['principal_card_key'].replace('nan', '')
    df['unique_card_key'] = df['unique_card_key'].replace('nan', '')
    
    dependent_rels = ['ابن', 'ابنة', 'بنت', 'زوج', 'زوجة', 'زوجه', 'أب', 'اب', 'أم', 'ام', 'أخ', 'اخ', 'أخت', 'اخت']
    df['is_principal'] = df.apply(lambda row: str(row['الصلة']).strip() not in dependent_rels, axis=1)
    
    # 2. Prepare data for template
    final_data = []
    for _, row in df.iterrows():
        final_data.append({
            'full_name': row['الاسم'],
            'employer': 'المنطقة الحرة جليانة' if row['is_principal'] else '',
            'principal_card_number': row['principal_card_key'] if not row['is_principal'] else '',
            'relationship': map_relationship(row['الصلة']),
            'card_number': row['unique_card_key'],
            'SORT_EMP': row['emp_name_norm'],
            'SORT_TYPE': 1 if row['is_principal'] else 0
        })
    
    final_df = pd.DataFrame(final_data)
    final_df = final_df.sort_values(by=['card_number', 'SORT_TYPE'], ascending=[True, False])
    final_df = final_df.drop_duplicates(subset=['card_number'])
    final_df = final_df.sort_values(by=['SORT_EMP', 'SORT_TYPE'], ascending=[True, False])
    
    # 3. Load Template and Fill
    template_path = r'd:\tba_waad_system-main\tba_waad_system-main\members_template (3).xlsx'
    output_path = r'd:\tba_waad_system-main\tba_waad_system-main\جليانة.xlsx'
    
    wb = load_workbook(template_path)
    ws = wb['Data']
    
    # Set headers in row 1
    ws.cell(row=1, column=1, value='الاسم الكامل\nfull_name')
    ws.cell(row=1, column=2, value='جهة العمل\nemployer')
    ws.cell(row=1, column=3, value='رقم بطاقة الرئيسي\nprincipal_card_number')
    ws.cell(row=1, column=4, value='القرابة\nrelationship')
    ws.cell(row=1, column=5, value='رقم البطاقة\ncard_number') # Explicit header for detection

    # Clear existing data from row 2
    for r in range(2, ws.max_row + 1):
        for c in range(1, 6):
            ws.cell(row=r, column=c).value = None

    row_idx = 2
    for _, series in final_df.iterrows():
        ws.cell(row=row_idx, column=1, value=series['full_name'])
        ws.cell(row=row_idx, column=2, value=series['employer'])
        ws.cell(row=row_idx, column=3, value=series['principal_card_number'])
        ws.cell(row=row_idx, column=4, value=series['relationship'])
        ws.cell(row=row_idx, column=5, value=series['card_number'])
        row_idx += 1
        
    wb.save(output_path)
    print(f"Data filled successfully. Final rows: {len(final_df)}")
    print(f"Saved to: {output_path}")

except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)
