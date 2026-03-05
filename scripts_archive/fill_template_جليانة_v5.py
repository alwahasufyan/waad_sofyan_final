import pandas as pd
import sys
import os
from openpyxl import load_workbook

def normalize_name(name):
    if not isinstance(name, str): return ""
    name = str(name).strip()
    if name.lower() in ['الاسم', 'اسم الموظف', 'الرقم الوظيفي', 'nan', 'none', '']: 
        return ""
    # Normalize spaces
    return " ".join(name.split())

def map_relationship(rel):
    rel = str(rel).strip()
    if not rel or rel == 'nan' or rel == 'موظف' or rel == 'Principal' or rel == 'نفسه' or rel == 'None' or rel == 'الصلة':
        return ""
    
    mapping = {
        'ابن': 'SON',
        'ابنة': 'DAUGHTER',
        'بنت': 'DAUGHTER',
        'زوج': 'HUSBAND',
        'زوجة': 'WIFE',
        'زوجه': 'WIFE',
        'أب': 'FATHER',
        'اب': 'FATHER',
        'أم': 'MOTHER',
        'ام': 'MOTHER',
        'أخ': 'BROTHER',
        'اخ': 'BROTHER',
        'أخت': 'SISTER',
        'اخت': 'SISTER'
    }
    return mapping.get(rel, rel)

try:
    source_file = r'd:\tba_waad_system-main\tba_waad_system-main\القائمة_الموحدة_جليانة_جاهزة_للرفع_دقيقة_100.xlsx'
    df = pd.read_excel(source_file)
    print(f"Reading source... total rows: {len(df)}")
    
    # 1. Standardize columns
    df['الاسم_نرم'] = df['الاسم'].apply(normalize_name)
    df['الموظف_نرم'] = df['اسم_الموظف'].apply(normalize_name) # Assuming this is parent name
    
    # Identify principals in source
    dep_rels = ['ابن', 'ابنة', 'بنت', 'زوج', 'زوجة', 'زوجه', 'أب', 'اب', 'أم', 'ام', 'أخ', 'اخ', 'أخت', 'اخت']
    df['is_src_principal'] = df.apply(lambda row: str(row['الصلة']).strip() not in dep_rels, axis=1)
    
    # Card numbers
    df['card_clean'] = df['رقم_بطاقة_المعالج'].astype(str).str.replace('.0', '', regex=False).str.strip().replace('nan', '')
    df['parent_card_clean'] = df['رقم_البطاقة_التأمينية'].astype(str).str.replace('.0', '', regex=False).str.strip().replace('nan', '')
    
    # 2. Build Principal Lookup (by Name)
    principal_lookup = {}
    for _, row in df[df['is_src_principal']].iterrows():
        if row['الاسم_نرم']:
            principal_lookup[row['الاسم_نرم']] = row['card_clean']
            
    print(f"Built principal lookup with {len(principal_lookup)} unique principals.")

    # 3. Process Data
    final_data = []
    skipped_rows = 0
    
    for _, row in df.iterrows():
        if not row['الاسم_نرم']:
            skipped_rows += 1
            continue
            
        is_principal = row['is_src_principal']
        rel = map_relationship(row['الصلة'])
        
        # If it's a dependent, try to find parent card if missing
        parent_card = row['parent_card_clean']
        if not is_principal and not parent_card:
            # Try to lookup by employee name
            parent_card = principal_lookup.get(row['الموظف_نرم'], "")
            
        final_data.append({
            'full_name': row['الاسم_نرم'],
            'employer': 'المنطقة الحرة جليانة' if is_principal else '',
            'principal_card_number': parent_card if not is_principal else '',
            'relationship': rel,
            'card_number': row['card_clean'],
            'SORT_EMP': row['الموظف_نرم'],
            'SORT_TYPE': 1 if is_principal else 0
        })
        
    final_df = pd.DataFrame(final_data)
    print(f"Skipped {skipped_rows} empty rows.")
    
    # Deduplicate: preferring principal if same card number exists
    final_df = final_df.sort_values(by=['card_number', 'SORT_TYPE'], ascending=[True, False])
    final_df = final_df.drop_duplicates(subset=['card_number'])
    
    # Global sort for better UX
    final_df = final_df.sort_values(by=['SORT_EMP', 'SORT_TYPE'], ascending=[True, False])
    
    # 4. Fill Template
    template_path = r'd:\tba_waad_system-main\tba_waad_system-main\members_template (3).xlsx'
    output_path = r'd:\tba_waad_system-main\tba_waad_system-main\جليانة.xlsx'
    
    wb = load_workbook(template_path)
    ws = wb['Data']
    
    # Set headers in row 1
    # System looks for these exact strings (case insensitive or parts)
    ws.cell(row=1, column=1, value='الاسم الكامل\nfull_name')
    ws.cell(row=1, column=2, value='جهة العمل\nemployer')
    ws.cell(row=1, column=3, value='رقم بطاقة الرئيسي\nprincipal_card_number')
    ws.cell(row=1, column=4, value='القرابة\nrelationship')
    ws.cell(row=1, column=5, value='رقم البطاقة\ncard_number')

    # Clear worksheet from row 2
    if ws.max_row >= 2:
        for r in range(2, ws.max_row + 1):
            for c in range(1, 10):
                ws.cell(row=r, column=c).value = None

    row_idx = 2
    for _, series in final_df.iterrows():
        ws.cell(row=row_idx, column=1, value=series['full_name'])
        ws.cell(row=row_idx, column=2, value=series['employer'])
        ws.cell(row=row_idx, column=3, value=series['principal_card_number'])
        ws.cell(row=row_idx, column=4, value=series['relationship'])
        ws.cell(row=row_idx, column=5, value=series['card_number'])
        row_idx += 1
        
    wb.save(output_path)
    print(f"SUCCESS: جليانة.xlsx generated with {len(final_df)} rows.")
    print(f"Output saved to: {output_path}")

except Exception as e:
    import traceback
    traceback.print_exc()
    sys.exit(1)
