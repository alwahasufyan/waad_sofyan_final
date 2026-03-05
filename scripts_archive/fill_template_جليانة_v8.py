import pandas as pd
import sys
import os
from openpyxl import load_workbook

def normalize_name(name):
    if not isinstance(name, str): return ""
    name = str(name).strip()
    if name.lower() in ['الاسم', 'اسم الموظف', 'الرقم الوظيفي', 'nan', 'none', '']: 
        return ""
    return " ".join(name.split())

def map_relationship(rel):
    rel = str(rel).strip()
    if not rel or rel == 'nan' or rel == 'موظف' or rel == 'Principal' or rel == 'نفسه' or rel == 'None' or rel == 'الصلة':
        return ""
    
    mapping = {
        'ابن': 'SON',
        'ابنة': 'DAUGHTER',
        'بنت': 'DAUGHTER',
        'زوج': 'HUSBAND',
        'زوجة': 'WIFE',
        'زوجه': 'WIFE',
        'أب': 'FATHER',
        'اب': 'FATHER',
        'أم': 'MOTHER',
        'ام': 'MOTHER',
        'أخ': 'BROTHER',
        'اخ': 'BROTHER',
        'أخت': 'SISTER',
        'اخت': 'SISTER'
    }
    return mapping.get(rel, rel)

try:
    source_file = r'd:\tba_waad_system-main\tba_waad_system-main\القائمة_الموحدة_جليانة_جاهزة_للرفع_دقيقة_100.xlsx'
    df = pd.read_excel(source_file)
    print(f"Reading source... total rows: {len(df)}")
    
    df['الاسم_نرم'] = df['الاسم'].apply(normalize_name)
    df['الموظف_نرم'] = df['اسم_الموظف'].apply(normalize_name)
    
    dep_rels = ['ابن', 'ابنة', 'بنت', 'زوج', 'زوجة', 'زوجه', 'أب', 'اب', 'أم', 'ام', 'أخ', 'اخ', 'أخت', 'اخت']
    df['is_src_principal'] = df.apply(lambda row: str(row['الصلة']).strip() not in dep_rels, axis=1)
    
    # Force uppercase for card numbers
    df['card_clean'] = df['رقم_بطاقة_المعالج'].astype(str).str.replace('.0', '', regex=False).str.strip().replace('nan', '').str.upper()
    df['parent_card_clean'] = df['رقم_البطاقة_التأمينية'].astype(str).str.replace('.0', '', regex=False).str.strip().replace('nan', '').str.upper()
    
    principal_lookup = {}
    for _, row in df[df['is_src_principal']].iterrows():
        if row['الاسم_نرم']:
            principal_lookup[row['الاسم_نرم']] = row['card_clean']
            
    final_data = []
    for _, row in df.iterrows():
        if not row['الاسم_نرم']: continue
            
        is_principal = row['is_src_principal']
        rel = map_relationship(row['الصلة'])
        parent_card = row['parent_card_clean']
        if not is_principal and not parent_card:
            parent_card = principal_lookup.get(row['الموظف_نرم'], "")
            
        final_data.append({
            'full_name': row['الاسم_نرم'],
            'employer': 'المنطقة الحرة جليانة' if is_principal else '',
            'principal_card_number': parent_card if not is_principal else '',
            'relationship': rel,
            'card_number': row['card_clean'],
            'SORT_EMP': row['الموظف_نرم'],
            'SORT_TYPE': 1 if is_principal else 0
        })
        
    final_df = pd.DataFrame(final_data)
    # Deduplicate: preferring principal record if multiple same card_number
    final_df = final_df.sort_values(by=['card_number', 'SORT_TYPE'], ascending=[True, False])
    final_df = final_df.drop_duplicates(subset=['card_number'])
    # Global sort to group families and put principal first
    final_df = final_df.sort_values(by=['SORT_EMP', 'SORT_TYPE'], ascending=[True, False])
    
    output_path = r'd:\tba_waad_system-main\tba_waad_system-main\جليانة.xlsx'
    template_path = r'd:\tba_waad_system-main\tba_waad_system-main\members_template (3).xlsx'
    
    # Header mapping to EXACT keywords in Java Pass 1
    column_mapping = {
        'full_name': 'الاسم الكامل',
        'employer': 'جهة العمل',
        'principal_card_number': 'رقم بطاقة الرئيسي',
        'relationship': 'القرابة',
        'card_number': 'رقم البطاقة'
    }
    
    wb = load_workbook(template_path)
    ws = wb['Data']
    
    # Set headers in row 1
    headers = list(column_mapping.values())
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)

    # Clear worksheet and set to row 3 (since firstDataRow = 2 usually means index 2 but in POI it is 0-indexed which is row 3 in Excel)
    # Wait, POI Row 0 is Excel 1. Row 1 is Excel 2.
    # In my code: firstDataRow = 2 (index 2) means Row index 2 = Excel Row 3.
    # So Row 1 (index 0) = Headers. Row 2 (index 1) = Example/Trash. Row 3 (index 2) = Start of data.
    
    # Clear and fill
    for r in ws.iter_rows(min_row=2):
        for cell in r:
            cell.value = None

    row_idx = 3 # Start data from Excel row 3
    for _, series in final_df.iterrows():
        ws.cell(row=row_idx, column=1, value=series['full_name'])
        ws.cell(row=row_idx, column=2, value=series['employer'])
        ws.cell(row=row_idx, column=3, value=series['principal_card_number'])
        ws.cell(row=row_idx, column=4, value=series['relationship'])
        ws.cell(row=row_idx, column=5, value=series['card_number'])
        row_idx += 1
        
    wb.save(output_path)
    print(f"SUCCESS: جليانة.xlsx V8 generated with {len(final_df)} rows. Card numbers are UPPERCASE.")

except Exception as e:
    import traceback
    traceback.print_exc()
    sys.exit(1)
