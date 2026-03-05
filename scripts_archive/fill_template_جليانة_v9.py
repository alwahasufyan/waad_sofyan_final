import pandas as pd
import sys
import os
from openpyxl import load_workbook

def normalize_name(name):
    if not isinstance(name, str): return ""
    name = str(name).strip()
    if name.lower() in ['الاسم', 'اسم الموظف', 'الرقم الوظيفي', 'nan', 'none', '']: 
        return ""
    return " ".join(name.split())

def map_relationship(rel):
    rel = str(rel).strip()
    if not rel or rel == 'nan' or rel == 'موظف' or rel == 'Principal' or rel == 'نفسه' or rel == 'None' or rel == 'الصلة':
        return ""
    
    mapping = {
        'ابن': 'SON',
        'ابنة': 'DAUGHTER',
        'بنت': 'DAUGHTER',
        'زوج': 'HUSBAND',
        'زوجة': 'WIFE',
        'زوجه': 'WIFE',
        'أب': 'FATHER',
        'اب': 'FATHER',
        'أم': 'MOTHER',
        'ام': 'MOTHER',
        'أخ': 'BROTHER',
        'اخ': 'BROTHER',
        'أخت': 'SISTER',
        'اخت': 'SISTER'
    }
    return mapping.get(rel, rel)

try:
    source_file = r'd:\tba_waad_system-main\tba_waad_system-main\القائمة_الموحدة_جليانة_جاهزة_للرفع_دقيقة_100.xlsx'
    df = pd.read_excel(source_file)
    print(f"Original source rows: {len(df)}")
    
    # Standardize names and cards
    df['الاسم_نرم'] = df['الاسم'].apply(normalize_name)
    df['اسم_الموظفة_نرم'] = df['اسم_الموظف'].apply(normalize_name)
    df['card_clean'] = df['رقم_بطاقة_المعالج'].astype(str).str.replace('.0', '', regex=False).str.strip().str.replace('nan', '', case=False).str.upper()
    df['parent_card_clean'] = df['رقم_البطاقة_التأمينية'].astype(str).str.replace('.0', '', regex=False).str.strip().str.replace('nan', '', case=False).str.upper()
    
    # Identify unique families by Parent Card
    unique_parent_cards = df['parent_card_clean'][df['parent_card_clean'].str.len() > 3].unique()
    print(f"Detected {len(unique_parent_cards)} unique families (parent cards).")
    
    # Check which families are missing their principal row
    existing_principals = df['card_clean'][df['card_clean'].isin(unique_parent_cards)].unique()
    missing_principals = set(unique_parent_cards) - set(existing_principals)
    print(f"Found {len(missing_principals)} families missing their principal row in source.")
    
    # Map each parent_card to its Employee Name for creating missing rows
    parent_to_emp_name = {}
    for _, row in df.iterrows():
        p_card = row['parent_card_clean']
        e_name = row['اسم_الموظفة_نرم']
        if p_card and e_name and p_card not in parent_to_emp_name:
            parent_to_emp_name[p_card] = e_name

    # 1. Start building final data with existing rows
    final_data = []
    processed_cards = set()
    
    # Add existing rows
    for _, row in df.iterrows():
        name = row['الاسم_نرم']
        if not name: continue
        
        card = row['card_clean']
        parent = row['parent_card_clean']
        is_p = (card == parent) and (card != "")
        
        rel = map_relationship(row['الصلة'])
        
        final_data.append({
            'full_name': name,
            'employer': 'المنطقة الحرة جليانة' if is_p else '',
            'principal_card_number': parent if not is_p else '',
            'relationship': rel,
            'card_number': card,
            'SORT_ORDER': 0 if is_p else 1,
            'SORT_GROUP': parent if parent else card
        })
        processed_cards.add(card)

    # 2. PROACTIVELY ADD MISSING PRINCIPALS
    added_count = 0
    for m_card in missing_principals:
        p_name = parent_to_emp_name.get(m_card, f"موظف {m_card}")
        
        final_data.append({
            'full_name': p_name,
            'employer': 'المنطقة الحرة جليانة',
            'principal_card_number': '',
            'relationship': '',
            'card_number': m_card,
            'SORT_ORDER': 0,
            'SORT_GROUP': m_card
        })
        added_count += 1
        
    print(f"Added {added_count} missing principal rows total.")
    
    # 3. Final cleanup and sorting
    final_df = pd.DataFrame(final_data)
    # Deduplicate by card (sometimes source has duplicates)
    final_df = final_df.drop_duplicates(subset=['card_number'])
    
    # Important sort: Group by family, Principal (SORT_ORDER 0) must be first in group
    final_df = final_df.sort_values(by=['SORT_GROUP', 'SORT_ORDER'], ascending=[True, True])
    
    print(f"Total Rows for output: {len(final_df)}")
    
    # 4. Fill Template
    output_path = r'd:\tba_waad_system-main\tba_waad_system-main\جليانة.xlsx'
    template_path = r'd:\tba_waad_system-main\tba_waad_system-main\members_template (3).xlsx'
    
    wb = load_workbook(template_path)
    ws = wb['Data']
    
    # Column mapping (Keywords for Java)
    column_mapping = {
        'full_name': 'الاسم الكامل',
        'employer': 'جهة العمل',
        'principal_card_number': 'رقم بطاقة الرئيسي',
        'relationship': 'القرابة',
        'card_number': 'رقم البطاقة'
    }
    
    # Set headers
    headers = list(column_mapping.values())
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)

    # Clear and fill
    for r in ws.iter_rows(min_row=2):
        for cell in r:
            cell.value = None

    row_idx = 3 # Data starts on Row 3 (firstDataRow=2 index)
    for _, series in final_df.iterrows():
        ws.cell(row=row_idx, column=1, value=series['full_name'])
        ws.cell(row=row_idx, column=2, value=series['employer'])
        ws.cell(row=row_idx, column=3, value=series['principal_card_number'])
        ws.cell(row=row_idx, column=4, value=series['relationship'])
        ws.cell(row=row_idx, column=5, value=series['card_number'])
        row_idx += 1
        
    wb.save(output_path)
    print(f"SUCCESS: جليانة.xlsx V9 generated with {len(final_df)} rows (including added principals).")

except Exception as e:
    import traceback
    traceback.print_exc()
    sys.exit(1)
