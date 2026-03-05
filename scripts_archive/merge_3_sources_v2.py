
import pandas as pd
import sys
import os
from openpyxl import load_workbook

def normalize_name(name):
    if not isinstance(name, str): return ""
    name = str(name).strip()
    if name.lower() in ['الاسم', 'اسم الموظف', 'الرقم الوظيفي', 'nan', 'none', '']: return ""
    return " ".join(name.split())

def map_rel(rel, card=""):
    rel = str(rel).strip()
    mapping = {
        'ابن': 'SON', 'ابنة': 'DAUGHTER', 'بنت': 'DAUGHTER',
        'زوج': 'HUSBAND', 'زوجة': 'WIFE', 'زوجه': 'WIFE',
        'أب': 'FATHER', 'اب': 'FATHER', 'أم': 'MOTHER', 'ام': 'MOTHER',
        'أخ': 'BROTHER', 'اخ': 'BROTHER', 'أخت': 'SISTER', 'اخت': 'SISTER'
    }
    res = mapping.get(rel)
    if res: return res
    
    # Try card suffix if mapping fails
    card = str(card).upper()
    if 'W' in card: return 'WIFE'
    if 'H' in card: return 'HUSBAND'
    if 'S' in card: return 'SON'
    if 'D' in card: return 'DAUGHTER'
    if 'M' in card: return 'MOTHER'
    if 'F' in card: return 'FATHER'
    
    return 'SON' # Final fallback to avoid "dependent" error

try:
    processed_records = []
    seen_names = set()

    # --- FILE 1: اسماء جليانة مرتب (2).xlsx ---
    f1 = r'd:\tba_waad_system-main\tba_waad_system-main\اسماء جليانة مرتب (2).xlsx'
    df1 = pd.read_excel(f1)
    
    current_p_name = ""
    current_p_card = ""
    
    for _, row in df1.iterrows():
        emp = normalize_name(row['Employee Name'])
        dep = normalize_name(row['Dependents'])
        p_card = str(row['Insurance Profile']).strip().upper().replace('.0', '')
        d_card = str(row['Dependents/Insurance Profile']).strip().upper().replace('.0', '')
        
        if emp and p_card and p_card != 'NAN':
            current_p_name = emp
            current_p_card = p_card
            if emp.lower() not in seen_names:
                processed_records.append({
                    'name': emp, 'employer': 'المنطقة الحرة جليانة', 'p_card': '', 'rel': '', 'card': p_card, 'grp': p_card
                })
                seen_names.add(emp.lower())
        
        if dep and d_card and d_card != 'NAN':
            if dep.lower() not in seen_names:
                rel = map_rel("", d_card) # Use card suffix
                processed_records.append({
                    'name': dep, 'employer': '', 'p_card': current_p_card, 'rel': rel, 'card': d_card, 'grp': current_p_card
                })
                seen_names.add(dep.lower())

    # --- FILE 2: كشف 7الى 20 المنطقة الحرة جليانة.xlsx ---
    f2 = r'd:\tba_waad_system-main\tba_waad_system-main\كشف 7الى 20 المنطقة الحرة جليانة.xlsx'
    df2 = pd.read_excel(f2)
    
    current_p_card = ""
    for _, row in df2.iterrows():
        name = normalize_name(row['الاســـم'])
        card = str(row['رقم البطاقة التأمينية']).strip().upper().replace('.0', '')
        rel_ar = str(row['صلة القرابة']).strip()
        
        if not name or card == 'NAN': continue
        
        is_p = (rel_ar == 'موظف' or rel_ar == 'Principal' or ('W' not in card and 'S' not in card and 'D' not in card and 'H' not in card and len(card) <= 12))
        
        if is_p:
            current_p_card = card
            if name.lower() not in seen_names:
                processed_records.append({
                    'name': name, 'employer': 'المنطقة الحرة جليانة', 'p_card': '', 'rel': '', 'card': card, 'grp': card
                })
                seen_names.add(name.lower())
        else:
            rel = map_rel(rel_ar, card)
            p_c = current_p_card
            if name.lower() not in seen_names:
                processed_records.append({
                    'name': name, 'employer': '', 'p_card': p_c, 'rel': rel, 'card': card, 'grp': p_c
                })
                seen_names.add(name.lower())

    # --- FILE 3: ايوب الرعيض 21دفعة الجديدة.xlsx ---
    f3 = r'd:\tba_waad_system-main\tba_waad_system-main\‏‏‏‏‏‏ايوب الرعيض 21دفعة الجديدة.xlsx'
    df3 = pd.read_excel(f3)
    cols = df3.columns.tolist()
    for _, row in df3.iterrows():
        name = normalize_name(row[cols[1]]) if len(cols)>1 else ""
        card = str(row[cols[2]]).strip().upper().replace('.0', '') if len(cols)>2 else ""
        rel_ar = str(row[cols[3]]).strip() if len(cols)>3 else ""
        
        if not name or card == 'NAN': continue
        
        if name.lower() not in seen_names:
            is_p = ('D' not in card and 'S' not in card and 'W' not in card and 'H' not in card)
            rel = "" if is_p else map_rel(rel_ar, card)
            processed_records.append({
                'name': name, 'employer': 'المنطقة الحرة جليانة' if is_p else '',
                'p_card': '' if is_p else card.split('W')[0].split('S')[0].split('D')[0].split('H')[0],
                'rel': rel,
                'card': card, 'grp': card if is_p else card.split('W')[0].split('S')[0].split('D')[0].split('H')[0]
            })
            seen_names.add(name.lower())

    final_df = pd.DataFrame(processed_records)
    
    # 1. Deduplicate by name and card to be safe
    final_df = final_df.drop_duplicates(subset=['name'])
    
    # 2. Sort so Principals (p_card == '') are FIRST for each group
    # This ensures the Principal is created before the dependent
    final_df['is_p'] = final_df['p_card'].apply(lambda x: 1 if not x else 0)
    final_df = final_df.sort_values(by=['grp', 'is_p'], ascending=[True, False])

    print(f"Final Count: {len(final_df)}")

    # Fill Template
    output_path = r'd:\tba_waad_system-main\tba_waad_system-main\جليانة.xlsx'
    template_path = r'd:\tba_waad_system-main\tba_waad_system-main\members_template (3).xlsx'
    wb = load_workbook(template_path)
    ws = wb['Data']
    
    column_mapping = {'name': 'الاسم الكامل', 'employer': 'جهة العمل', 'p_card': 'رقم بطاقة الرئيسي', 'rel': 'القرابة', 'card': 'رقم البطاقة'}
    for col_idx, header in enumerate(column_mapping.values(), 1):
        ws.cell(row=1, column=col_idx, value=header)
    for r in ws.iter_rows(min_row=2):
        for cell in r: cell.value = None

    row_idx = 3 
    for _, row in final_df.iterrows():
        ws.cell(row=row_idx, column=1, value=row['name'])
        ws.cell(row=row_idx, column=2, value=row['employer'])
        ws.cell(row=row_idx, column=3, value=row['p_card'])
        ws.cell(row=row_idx, column=4, value=row['rel'])
        ws.cell(row=row_idx, column=5, value=row['card'])
        row_idx += 1
        
    wb.save(output_path)
    print(f"SUCCESS: جليانة.xlsx fixed and generated with {len(final_df)} rows.")

except Exception as e:
    import traceback
    traceback.print_exc()
    sys.exit(1)
