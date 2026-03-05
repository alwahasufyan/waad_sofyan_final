
import pandas as pd
import re
import sys
from openpyxl import load_workbook

def normalize_name(name):
    if pd.isna(name): return ""
    name = str(name).strip()
    if name.lower() in ['الاسم', 'اسم الموظف', 'الرقم الوظيفي', 'nan', 'none', '*', '']: return ""
    return " ".join(name.split())

def extract_fnum(val):
    if pd.isna(val): return None
    s = str(val).upper().replace('.0', '').strip()
    match = re.search(r'JFZ(?:2025)?(0?\d{4,})', s)
    if match: return match.group(1).lstrip('0')
    digits = ''.join(filter(str.isdigit, s))
    if len(digits) > 2: return digits.lstrip('0')
    return None

def standardize_rel(rel_str):
    s = str(rel_str).strip().lower()
    if s in ['الموظف', 'موظف', 'principal', 'رئيسي']: return 'PRINCIPAL'
    mapping = {
        'ابن': 'SON', 'ابنة': 'DAUGHTER', 'بنت': 'DAUGHTER',
        'زوج': 'HUSBAND', 'زوجة': 'WIFE', 'زوجه': 'WIFE', 'الزوجة': 'WIFE',
        'أب': 'FATHER', 'اب': 'FATHER', 'أم': 'MOTHER', 'ام': 'MOTHER',
        'أخ': 'BROTHER', 'اخ': 'BROTHER', 'أخت': 'SISTER', 'اخت': 'SISTER'
    }
    for k, v in mapping.items():
        if k in s: return v
    return 'DEPENDENT'

def parse_suffix(card):
    # Extracts the letter and number suffix if any -> ('S', 1)
    match = re.search(r'([A-Z])(\d+)$', str(card).upper())
    if match: return match.group(1), int(match.group(2))
    return None, 1

families = {}
# families[f_num] = { 'members': [ { name, rel, raw_card, final_card, is_p } ] }

def add_to_family(f_num, name, rel, raw_card):
    if not f_num or not name: return
    if f_num not in families: families[f_num] = []
    
    # Check strict uniqueness by name
    for m in families[f_num]:
        if m['name'] == name: return # skip Duplicate across files!
        
    families[f_num].append({
        'name': name,
        'rel': standardize_rel(rel) if pd.notna(rel) and str(rel)!='nan' else 'DEPENDENT',
        'raw_card': str(raw_card).upper().strip().replace('.0', '') if pd.notna(raw_card) else ''
    })

# --- FILE 1 ---
f1 = r'd:\tba_waad_system-main\tba_waad_system-main\اسماء جليانة مرتب (2).xlsx'
df1 = pd.read_excel(f1)
for _, row in df1.iterrows():
    emp = normalize_name(row['Employee Name'])
    dep = normalize_name(row['Dependents'])
    p_card = str(row['Insurance Profile'])
    d_card = str(row['Dependents/Insurance Profile'])
    
    f_num = extract_fnum(p_card) or extract_fnum(row['Sequence'])
    if not f_num: continue
    
    if emp:
        add_to_family(f_num, emp, 'PRINCIPAL', p_card)
    if dep:
        add_to_family(f_num, dep, 'DEPENDENT', d_card)

# --- FILE 2 ---
f2 = r'd:\tba_waad_system-main\tba_waad_system-main\كشف 7الى 20 المنطقة الحرة جليانة.xlsx'
df2 = pd.read_excel(f2)
for _, row in df2.iterrows():
    name = normalize_name(row['الاســـم'])
    raw_card = row['رقم البطاقة التأمينية']
    rel = row['صلة القرابة']
    
    f_num = extract_fnum(raw_card)
    if not f_num: continue
    
    # If it's a raw number but the relationship isn't PRINCIPAL, standardize_rel handles the rest
    # But wait, we should pass it to add_to_family and let standardizer deal with it
    add_to_family(f_num, name, rel, raw_card)

# --- FILE 3 ---
f3 = r'd:\tba_waad_system-main\tba_waad_system-main\‏‏‏‏‏‏ايوب الرعيض 21دفعة الجديدة.xlsx'
df3 = pd.read_excel(f3)
df3['الرقم المالي '] = df3['الرقم المالي '].ffill()
cols = df3.columns.tolist()
for _, row in df3.iterrows():
    name = normalize_name(row[cols[1]]) if len(cols)>1 else ""
    raw_card = row['Unnamed: 4']
    rel = row['الصلة']
    f_num_raw = row['الرقم المالي ']
    
    f_num = extract_fnum(f_num_raw) or extract_fnum(raw_card)
    if not f_num: continue
    add_to_family(f_num, name, rel, raw_card)

# --- Post-process and Assign Final Cards ---
processed_records = []
for f_num, members in families.items():
    if not members: continue
    
    # Identify Principal
    # It's either marked PRINCIPAL, or the one without Suffix, or just the first if ambiguous
    p_member = None
    for m in members:
        if m['rel'] == 'PRINCIPAL':
            p_member = m
            break
            
    if not p_member:
        # Try to find someone whose card has no suffix
        for m in members:
            # If card is something like '31060' or 'JFZ202531060' (len <= 12)
            if 'S' not in m['raw_card'] and 'W' not in m['raw_card'] and 'D' not in m['raw_card'] and 'H' not in m['raw_card']:
                p_member = m
                break
                
    if not p_member:
        p_member = members[0] # Fallback
    
    # Force Principal status
    p_member['rel'] = 'PRINCIPAL'
    p_member['final_card'] = f"JFZ2025{f_num}"
    
    # Now assign to dependents
    # Count existing suffixes to avoid duplicate suffix (like assigning S1 twice)
    counters = {'W':0, 'S':0, 'D':0, 'H':0, 'M':0, 'F':0, 'B':0, 'C':0}
    
    # First pass: find existing valid JFZ suffixes and update counters
    for m in members:
        if m == p_member: continue
        if 'JFZ' in m['raw_card']:
            l, num = parse_suffix(m['raw_card'])
            if l and l in counters:
                counters[l] = max(counters[l], num)
                
    # Second pass: assign final cards
    for m in members:
        if m == p_member: continue
        
        # If it's a valid JFZ card WITH suffix, keep it
        if 'JFZ' in m['raw_card'] and any(char.isalpha() for char in m['raw_card'].replace('JFZ', '')):
            m['final_card'] = m['raw_card']
        else:
            # Generate one
            letter = 'S'
            if m['rel'] == 'WIFE': letter = 'W'
            elif m['rel'] == 'SON': letter = 'S'
            elif m['rel'] == 'DAUGHTER': letter = 'D'
            elif m['rel'] == 'HUSBAND': letter = 'H'
            elif m['rel'] == 'MOTHER': letter = 'M'
            elif m['rel'] == 'FATHER': letter = 'F'
            elif m['rel'] == 'SISTER': letter = 'D' # fallback
            elif m['rel'] == 'BROTHER': letter = 'S' # fallback
            elif 'زوج' in m['raw_card']: letter = 'W' # try to infer from raw data if any
            
            counters[letter] += 1
            m['final_card'] = f"JFZ2025{f_num}{letter}{counters[letter]}"

    # Add to final flat list
    for m in members:
        is_p = (m == p_member)
        processed_records.append({
            'name': m['name'],
            'employer': 'المنطقة الحرة جليانة' if is_p else '',
            'p_card': '' if is_p else p_member['final_card'],
            'rel': '' if is_p else m['rel'],
            'card': m['final_card'],
            'f_num': f_num,
            'is_p': is_p
        })

# Create DataFrame
final_df = pd.DataFrame(processed_records)

# Cleanup relations, generic 'DEPENDENT' to SON for safety if still dangling
final_df['rel'] = final_df['rel'].replace('DEPENDENT', 'SON')

# Final Sort: Principal First, then Dependents
final_df = final_df.sort_values(by=['f_num', 'is_p'], ascending=[True, False])

print(f"Total rows after deep merge: {len(final_df)}")

# Fill Template
output_path = r'd:\tba_waad_system-main\tba_waad_system-main\جليانة.xlsx'
template_path = r'd:\tba_waad_system-main\tba_waad_system-main\members_template (3).xlsx'
wb = load_workbook(template_path)
ws = wb['Data']

column_mapping = {'name': 'الاسم الكامل', 'employer': 'جهة العمل', 'p_card': 'رقم بطاقة الرئيسي', 'rel': 'القرابة', 'card': 'رقم البطاقة'}
for col_idx, header in enumerate(column_mapping.values(), 1):
    ws.cell(row=1, column=col_idx, value=header)
for r in ws.iter_rows(min_row=2):
    for cell in r: cell.value = None

row_idx = 3 
for _, row in final_df.iterrows():
    ws.cell(row=row_idx, column=1, value=row['name'])
    ws.cell(row=row_idx, column=2, value=row['employer'])
    ws.cell(row=row_idx, column=3, value=row['p_card'])
    ws.cell(row=row_idx, column=4, value=row['rel'])
    ws.cell(row=row_idx, column=5, value=row['card'])
    row_idx += 1
    
wb.save(output_path)
print(f"SUCCESS: جليانة.xlsx generated accurately with {len(final_df)} rows based purely on Financial Numbers.")

