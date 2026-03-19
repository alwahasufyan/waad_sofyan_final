import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


ROOT_INPATIENT = "الإيواء"
ROOT_OUTPATIENT = "العيادات الخارجية"

SUB_INPATIENT_GENERAL = "الإيواء - عام"
SUB_INPATIENT_HOME_NURSING = "الإيواء - تمريض منزلي"
SUB_INPATIENT_PHYSIO = "الإيواء - علاج طبيعي"
SUB_INPATIENT_WORK_INJURY = "الإيواء - إصابات عمل"
SUB_INPATIENT_PSYCH = "الإيواء - طب نفسي"
SUB_INPATIENT_DELIVERY = "الإيواء - ولادة طبيعية وقيصرية"
SUB_INPATIENT_PREGNANCY = "الإيواء - مضاعفات حمل"
SUB_OUTPATIENT_GENERAL = "العيادات الخارجية - عام"
SUB_OUTPATIENT_RAD = "العيادات الخارجية - أشعة"
SUB_OUTPATIENT_MRI = "العيادات الخارجية - رنين مغناطيسي"
SUB_OUTPATIENT_DRUGS = "العيادات الخارجية - علاجات وأدوية"
SUB_OUTPATIENT_DEVICES = "العيادات الخارجية - أجهزة ومعدات"
SUB_OUTPATIENT_PHYSIO = "العيادات الخارجية - علاج طبيعي"
SUB_OUTPATIENT_DENTAL_ROUTINE = "العيادات الخارجية - أسنان روتيني"
SUB_OUTPATIENT_DENTAL_COSMETIC = "العيادات الخارجية - أسنان تجميلي"
SUB_OUTPATIENT_GLASSES = "العيادات الخارجية - النظارة الطبية"


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"[\u064b-\u065f\u0670]", "", text)
    text = text.replace("أ", "ا").replace("إ", "ا").replace("آ", "ا")
    text = text.replace("ة", "ه")
    text = re.sub(r"\s+", " ", text)
    return text


def has_any(text: str, keywords: list[str]) -> bool:
    return any(keyword in text for keyword in keywords)


def extract_price(value: object) -> float | None:
    if value is None:
        return None
    cleaned = "".join(ch for ch in str(value) if ch.isdigit() or ch == ".")
    if not cleaned:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def extract_service_code(service_cell: str) -> tuple[str, str]:
    service_cell = service_cell.strip()
    match = re.match(r"^([A-Za-z][A-Za-z0-9\-/]{0,30})\s+(.+)$", service_cell)
    if not match:
        return "", service_cell
    return match.group(1).strip(), match.group(2).strip()


def classify_row(service_name: str, specialty: str, raw_category: str) -> tuple[str, str, str, str]:
    service_n = normalize_text(service_name)
    specialty_n = normalize_text(specialty)
    raw_n = normalize_text(raw_category)
    text = " ".join(part for part in [service_n, specialty_n, raw_n] if part)

    if has_any(text, ["تمريض منزلي", "رعايه منزليه", "home care", "nursing"]):
        return ROOT_INPATIENT, SUB_INPATIENT_HOME_NURSING, "high", "home-nursing-keywords"

    if has_any(text, ["اصابات العمل", "اصابه عمل", "حادث عمل", "work injury"]):
        return ROOT_INPATIENT, SUB_INPATIENT_WORK_INJURY, "high", "work-injury-keywords"

    if has_any(text, ["طب نفسي", "جلسه نفسيه", "نفسي", "psychi"]):
        return ROOT_INPATIENT, SUB_INPATIENT_PSYCH, "high", "psych-keywords"

    if has_any(text, ["مضاعفات الحمل", "نزيف حمل", "اجهاض", "الحمل عالي الخطوره"]):
        return ROOT_INPATIENT, SUB_INPATIENT_PREGNANCY, "high", "pregnancy-complication-keywords"

    if has_any(text, ["ولاده", "قيصريه", "توليد", "c-section", "cesarean"]):
        return ROOT_INPATIENT, SUB_INPATIENT_DELIVERY, "high", "delivery-keywords"

    if has_any(text, ["علاج طبيعي", "العلاج الطبيعي", "جلسه علاج", "تاهيل", "physio", "physiotherapy"]):
        if has_any(raw_n + " " + specialty_n, ["ايواء", "اقامه", "داخل المستشفى", "رعايه", "عنايه"]):
            return ROOT_INPATIENT, SUB_INPATIENT_PHYSIO, "medium", "physio-with-inpatient-context"
        return ROOT_OUTPATIENT, SUB_OUTPATIENT_PHYSIO, "high", "physio-keywords"

    if "خدمات العلاج الطبيعي" in specialty_n:
        return ROOT_OUTPATIENT, SUB_OUTPATIENT_PHYSIO, "high", "physio-specialty-context"

    if has_any(raw_n, ["اسنان تجميلي"]) or has_any(text, ["تقويم", "تركيبه", "تركيبات", "زراعه", "bridge", "crown", "veneer", "denture", "implant"]):
        return ROOT_OUTPATIENT, SUB_OUTPATIENT_DENTAL_COSMETIC, "high", "dental-cosmetic-keywords"

    if has_any(raw_n, ["اسنان وقائي"]) or ("خدمات الاسنان" in specialty_n):
        return ROOT_OUTPATIENT, SUB_OUTPATIENT_DENTAL_ROUTINE, "high", "dental-routine-context"

    if has_any(text, ["نظاره", "نظارات", "عدسه", "عدسات", "frame", "glasses"]):
        return ROOT_OUTPATIENT, SUB_OUTPATIENT_GLASSES, "high", "glasses-keywords"

    if has_any(raw_n, ["اشعه"]) or has_any(specialty_n, ["اشعه", "الصور التشخيصيه"]):
        if has_any(text, ["رنين", "مقطعي", "ct", "mri", "scan"]):
            return ROOT_OUTPATIENT, SUB_OUTPATIENT_MRI, "high", "advanced-imaging-keywords"
        return ROOT_OUTPATIENT, SUB_OUTPATIENT_RAD, "high", "radiology-context"

    if has_any(raw_n, ["تحاليل طبيه"]) or specialty_n == "معامل":
        return ROOT_OUTPATIENT, SUB_OUTPATIENT_GENERAL, "medium", "lab-mapped-to-general"

    if has_any(text, ["دواء", "ادويه", "مستلزمات", "صيدليه", "ampoule", "capsule"]):
        return ROOT_OUTPATIENT, SUB_OUTPATIENT_DRUGS, "medium", "drug-keywords"

    if has_any(text, ["جهاز", "قسطره", "انبوب", "جبس", "رباط", "ضماد", "دعامة", "stent"]):
        if has_any(raw_n + " " + specialty_n, ["ايواء", "عمليات"]):
            return ROOT_INPATIENT, SUB_INPATIENT_GENERAL, "medium", "device-in-inpatient-context"
        return ROOT_OUTPATIENT, SUB_OUTPATIENT_DEVICES, "medium", "device-keywords"

    if has_any(raw_n, ["ايواء", "عمليات"]) or has_any(specialty_n, ["التخذير", "الرعايه الطبيه", "العنايه المركزه", "جراحه", "مناظير", "العلاج الكيماوي", "جلسات الغسيل"]):
        return ROOT_INPATIENT, SUB_INPATIENT_GENERAL, "medium", "inpatient-or-procedure-context"

    if has_any(raw_n, ["عيادات خارجيه"]) or has_any(specialty_n, ["خدمات العيادات الخارجيه", "كشف", "مراجعه", "العظام", "الاذن والانف والحنجره", "العيون", "الجراحه", "جراحه التجميل"]):
        return ROOT_OUTPATIENT, SUB_OUTPATIENT_GENERAL, "medium", "outpatient-context"

    return ROOT_OUTPATIENT, SUB_OUTPATIENT_GENERAL, "low", "fallback-general"


def build_notes(raw_category: str, confidence: str, reason: str) -> str:
    raw_part = raw_category.strip() if raw_category else "بدون تصنيف خام"
    prefix = "مراجعة يدوية" if confidence == "low" else "مصنف آلياً"
    return f"{prefix} | المصدر: {raw_part} | confidence={confidence} | rule={reason}"


def read_source_rows(source_path: Path) -> list[dict[str, object]]:
    df = pd.read_excel(source_path, header=None)
    services: list[dict[str, object]] = []

    for row_number, row in enumerate(df.itertuples(index=False), start=1):
        price = extract_price(row[2] if len(row) > 2 else None)
        service_cell = str(row[3]).strip() if len(row) > 3 and pd.notnull(row[3]) else ""
        specialty = str(row[4]).strip() if len(row) > 4 and pd.notnull(row[4]) else ""
        raw_category = str(row[5]).strip() if len(row) > 5 and pd.notnull(row[5]) else ""

        if not service_cell or any(token in service_cell for token in ["الخدمه", "الخدمة", "اسم الخدمة"]):
            continue
        if price is None or price <= 0:
            continue

        service_code, service_name = extract_service_code(service_cell)
        main_category, sub_category, confidence, reason = classify_row(service_name, specialty, raw_category)

        if not service_code:
            service_code = f"DS-{len(services) + 1:05d}"

        services.append(
            {
                "source_row": row_number,
                "source_service_cell": service_cell,
                "service_name": service_name,
                "service_code": service_code,
                "standard_price": price,
                "contract_price": price,
                "main_category": main_category,
                "sub_category": sub_category,
                "specialty": specialty,
                "raw_category": raw_category,
                "confidence": confidence,
                "reason": reason,
                "notes": build_notes(raw_category, confidence, reason),
            }
        )

    return services


def write_workbook(template_path: Path, output_path: Path, services: list[dict[str, object]]) -> None:
    workbook = load_workbook(template_path)
    pricing_sheet = workbook["Pricing_Template"]

    if pricing_sheet.max_row >= 2:
        pricing_sheet.delete_rows(2, pricing_sheet.max_row - 1)

    for row_index, service in enumerate(services, start=2):
        pricing_sheet.cell(row=row_index, column=1, value=service["service_name"])
        pricing_sheet.cell(row=row_index, column=2, value=service["service_code"])
        pricing_sheet.cell(row=row_index, column=3, value=service["standard_price"])
        pricing_sheet.cell(row=row_index, column=4, value=service["contract_price"])
        pricing_sheet.cell(row=row_index, column=5, value=service["main_category"])
        pricing_sheet.cell(row=row_index, column=6, value=service["sub_category"])
        pricing_sheet.cell(row=row_index, column=7, value=service["specialty"])
        pricing_sheet.cell(row=row_index, column=8, value=service["notes"])

    if "Classification_Audit" in workbook.sheetnames:
        del workbook["Classification_Audit"]

    audit_sheet = workbook.create_sheet("Classification_Audit")
    audit_headers = [
        "source_row",
        "source_service_cell",
        "service_code",
        "service_name",
        "standard_price",
        "specialty",
        "raw_category",
        "main_category",
        "sub_category",
        "confidence",
        "reason",
        "notes",
    ]
    audit_sheet.append(audit_headers)

    for service in services:
        audit_sheet.append([service[column] for column in audit_headers])

    workbook.save(output_path)


def print_summary(services: list[dict[str, object]]) -> None:
    df = pd.DataFrame(services)
    print(f"✅ عدد الخدمات الجاهزة: {len(df)}")
    print("\n📊 التوزيع حسب التصنيف الرئيسي:")
    print(df.groupby("main_category")["service_code"].count().to_string())
    print("\n📊 التوزيع حسب التصنيف الفرعي:")
    print(df.groupby("sub_category")["service_code"].count().sort_values(ascending=False).to_string())
    print("\n📊 حالات المراجعة اليدوية:")
    print(df.groupby("confidence")["service_code"].count().to_string())


def main() -> int:
    if len(sys.argv) < 4:
        print("الاستخدام: python build_dar_shifa_contract_price_list.py <source.xlsx> <template.xlsx> <output.xlsx>")
        return 1

    source_path = Path(sys.argv[1])
    template_path = Path(sys.argv[2])
    output_path = Path(sys.argv[3])

    if not source_path.exists():
        print(f"❌ ملف المصدر غير موجود: {source_path}")
        return 1
    if not template_path.exists():
        print(f"❌ ملف القالب غير موجود: {template_path}")
        return 1

    services = read_source_rows(source_path)
    write_workbook(template_path, output_path, services)
    print_summary(services)
    print(f"\n📁 تم إنشاء الملف: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())